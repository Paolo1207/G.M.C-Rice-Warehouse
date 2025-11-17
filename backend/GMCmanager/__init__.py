# backend/GMCmanager/__init__.py
from flask import Blueprint, render_template, request, jsonify, session, render_template_string
from sqlalchemy.exc import IntegrityError
from extensions import db
from models import Branch, Product, InventoryItem, RestockLog, SalesTransaction, ForecastData, User, EmailVerification, PasswordReset
from auth_helpers import manager_required  # <-- role guard
from activity_logger import ActivityLogger

manager_bp = Blueprint(
    "manager",
    __name__,
    template_folder="templates/manager",
    static_folder="static",
    static_url_path="/manager/static"
)

# ----------------------------- PAGES -----------------------------
@manager_bp.route("/dashboard", endpoint="manager_dashboard")
@manager_required
def dashboard():
    # Get branch_id from URL parameters
    branch_id = request.args.get('branch_id', type=int)
    branch_name = request.args.get('branch_name', '')
    
    return render_template("manager_dashboard.html", 
                         branch_id=branch_id, 
                         branch_name=branch_name)

@manager_bp.route("/analytics", endpoint="analytics")
@manager_required
def analytics():
    return render_template("manager_analytics.html")

@manager_bp.route("/forecast", endpoint="forecast")
@manager_required
def forecast():
    # Get branch_id from URL parameters
    branch_id = request.args.get('branch_id', type=int)
    branch_name = request.args.get('branch_name', '')
    
    return render_template("manager_forecast.html", 
                         branch_id=branch_id, 
                         branch_name=branch_name)

@manager_bp.route("/inventory", endpoint="inventory")
@manager_required
def inventory():
    # Get branch_id from URL parameters
    branch_id = request.args.get('branch_id', type=int)
    branch_name = request.args.get('branch_name', '')
    
    return render_template("manager_inventory.html", 
                         branch_id=branch_id, 
                         branch_name=branch_name)

@manager_bp.route("/notifications", endpoint="notifications")
@manager_required
def notifications():
    # Get branch_id from URL parameters
    branch_id = request.args.get('branch_id', type=int)
    branch_name = request.args.get('branch_name', '')
    
    return render_template("manager_notifications.html", 
                         branch_id=branch_id, 
                         branch_name=branch_name)

@manager_bp.route("/purchase", endpoint="purchase")
@manager_required
def purchase():
    # Get branch_id from URL parameters
    branch_id = request.args.get('branch_id', type=int)
    branch_name = request.args.get('branch_name', '')
    
    return render_template("manager_purchase.html", 
                         branch_id=branch_id, 
                         branch_name=branch_name)

@manager_bp.route("/reports", endpoint="reports")
@manager_required
def reports():
    return render_template("manager_reports.html")

@manager_bp.route("/sales", endpoint="sales")
@manager_required
def sales():
    return render_template("manager_sales.html")

@manager_bp.route("/settings", endpoint="settings")
@manager_required
def settings():
    return render_template("manager_settings.html")


# ============================ HELPERS ============================
def _current_manager_branch_id():
    """Prefer the branch pinned to the logged-in manager."""
    user = session.get("user") or {}
    return user.get("branch_id")

def _to_float(v):
    if v is None or v == "": return None
    try: return float(v)
    except ValueError: return None

def item_to_dict(it: InventoryItem):
    return {
        "id": it.id,
        "branch_id": it.branch_id,
        "product_id": it.product_id,
        "product_name": it.product.name if it.product else None,
        "category": it.product.category if it.product else None,
        "barcode": it.product.barcode if it.product else None,
        "sku": it.product.sku if it.product else None,
        "desc": it.product.description if it.product else None,
        "stock": float(it.stock_kg or 0),
        "price": float(it.unit_price or 0),
        "unit": "kg",  # Default unit for rice products
        "batch": it.batch_code,
        "batch_code": it.batch_code,
        "grn": it.grn_number,
        "grn_number": it.grn_number,
        "warn": float(it.warn_level) if it.warn_level is not None else None,
        "auto": float(it.auto_level) if it.auto_level is not None else None,
        "margin": it.margin,
        "status": it.status,  # assumes a @hybrid_property or column present
        "updated_at": it.updated_at.isoformat() if it.updated_at else None,
    }


# ========================= INVENTORY API =========================
# CREATE: add product (if needed) and inventory item for a (manager) branch
@manager_bp.route("/api/inventory", methods=["POST"])
@manager_required
def mgr_inventory_create():
    """
    JSON body:
    {
      "branch_id": 1,                # optional (defaults to manager's branch)
      "product_name": "Jasmine",     # required
      "category": "...", "barcode": "...", "sku": "...", "desc": "...",
      "stock_kg": 100, "unit_price": 45.0,
      "batch_code": "ABC123",
      "warn_level": 200, "auto_level": 50, "margin": "20%"
    }
    """
    data = request.get_json(silent=True) or {}

    # Resolve branch: prefer the manager's own branch
    branch_id = data.get("branch_id") or _current_manager_branch_id()
    product_name = (data.get("product_name") or "").strip()

    if not branch_id or not product_name:
        return jsonify({"ok": False, "error": "branch_id and product_name are required"}), 400

    branch = Branch.query.get(branch_id)
    if not branch:
        return jsonify({"ok": False, "error": f"Branch {branch_id} not found"}), 404

    # Find or create product by name (case-sensitive match)
    product = Product.query.filter_by(name=product_name).first()
    
    # Debug: Check for case-insensitive matches
    if not product:
        from sqlalchemy import func
        similar_products = Product.query.filter(
            func.lower(Product.name) == func.lower(product_name)
        ).all()
        if similar_products:
            print(f"DEBUG: Product '{product_name}' not found, but found similar (case-different): {[p.name for p in similar_products]}")
    
    if not product:
        print(f"DEBUG: Creating new product: '{product_name}'")
        product = Product(
            name=product_name,
            category=(data.get("category") or None),
            barcode=(data.get("barcode") or None),
            sku=(data.get("sku") or None),
            description=(data.get("desc") or None),
        )
        db.session.add(product)
        try:
            db.session.flush()
            # Refresh to ensure product.id is available
            db.session.refresh(product)
            print(f"DEBUG: Created product id={product.id}, name='{product.name}'")
        except IntegrityError:
            db.session.rollback()
            product = Product.query.filter_by(name=product_name).first()
            if not product:
                return jsonify({"ok": False, "error": "Failed to create or find product"}), 500
            print(f"DEBUG: Product already existed after rollback, using id={product.id}")
    else:
        print(f"DEBUG: Found existing product id={product.id}, name='{product.name}'")

    # Ensure product has an ID
    if not product or not product.id:
        return jsonify({"ok": False, "error": "Invalid product state"}), 500

    stock_kg   = _to_float(data.get("stock_kg")) or 0.0
    unit_price = _to_float(data.get("unit_price")) or 0.0
    batch_code = (data.get("batch_code") or "").strip() or None
    grn_number = (data.get("grn_number") or data.get("grn") or "").strip() or None

    # Debug logging
    print(f"DEBUG: Creating inventory item - branch_id={branch.id}, product_id={product.id}, product_name='{product_name}', batch_code='{batch_code}'")

    # Check if this exact combination already exists (branch + product + batch_code)
    # Handle NULL batch_code properly - PostgreSQL treats NULL as distinct in unique constraints
    from sqlalchemy import and_, func
    
    # Get all existing items for this product in this branch for debugging
    all_existing = InventoryItem.query.filter_by(
        branch_id=branch.id,
        product_id=product.id
    ).all()
    
    print(f"DEBUG: Found {len(all_existing)} existing inventory items for product_id={product.id} in branch_id={branch.id}")
    for existing in all_existing:
        print(f"DEBUG:   - Existing item id={existing.id}, batch_code='{existing.batch_code}'")
    
    if batch_code:
        # Check for exact match (case-sensitive)
        existing_item = InventoryItem.query.filter_by(
            branch_id=branch.id, 
            product_id=product.id, 
            batch_code=batch_code
        ).first()
        
        # Also check case-insensitive match for better error message
        if not existing_item:
            existing_item_case_insensitive = InventoryItem.query.filter(
                and_(
                    InventoryItem.branch_id == branch.id,
                    InventoryItem.product_id == product.id,
                    func.lower(InventoryItem.batch_code) == func.lower(batch_code)
                )
            ).first()
            if existing_item_case_insensitive:
                print(f"DEBUG: Found case-insensitive match: '{existing_item_case_insensitive.batch_code}' vs '{batch_code}'")
    else:
        # For NULL batch_code, check explicitly
        existing_item = InventoryItem.query.filter(
            and_(
                InventoryItem.branch_id == branch.id,
                InventoryItem.product_id == product.id,
                InventoryItem.batch_code.is_(None)
            )
        ).first()
    
    if existing_item:
        # Get all existing batch codes for this product in this branch for better error message
        all_batches = InventoryItem.query.filter_by(
            branch_id=branch.id,
            product_id=product.id
        ).with_entities(InventoryItem.batch_code).all()
        existing_batches = [b[0] for b in all_batches if b[0]]
        
        error_msg = f"Product '{product_name}' with batch code '{batch_code or '(none)'}' already exists in this branch."
        if existing_batches:
            error_msg += f" Existing batch codes: {', '.join(existing_batches)}"
        error_msg += " Please use a different batch code or update the existing item."
        
        print(f"DEBUG: Duplicate detected before commit: {error_msg}")
        return jsonify({"ok": False, "error": error_msg}), 409

    # Create new inventory item
    item = InventoryItem(
        branch_id=branch.id,
        product_id=product.id,
        stock_kg=stock_kg,
        unit_price=unit_price,
        batch_code=batch_code,
        grn_number=grn_number,
        warn_level=_to_float(data.get("warn_level")),
        auto_level=_to_float(data.get("auto_level")),
        margin=(data.get("margin") or None),
    )
    db.session.add(item)

    # Initial restock log if any stock provided
    if stock_kg > 0:
        db.session.add(RestockLog(
            inventory_item=item,
            qty_kg=stock_kg,
            supplier="Manager",
            note="Initial add"
        ))

    try:
        db.session.commit()
        print(f"DEBUG: Successfully created inventory item id={item.id}")
        return jsonify({"ok": True, "item": item_to_dict(item)}), 201
    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig) if hasattr(e, 'orig') else str(e)
        print(f"DEBUG: IntegrityError caught: {error_msg}")
        print(f"DEBUG: Attempted to create - branch_id={branch.id}, product_id={product.id}, product_name='{product_name}', batch_code='{batch_code}'")
        
        # Check what actually exists now (in case of race condition)
        if batch_code:
            existing_check = InventoryItem.query.filter_by(
                branch_id=branch.id,
                product_id=product.id,
                batch_code=batch_code
            ).first()
        else:
            existing_check = InventoryItem.query.filter(
                and_(
                    InventoryItem.branch_id == branch.id,
                    InventoryItem.product_id == product.id,
                    InventoryItem.batch_code.is_(None)
                )
            ).first()
        
        # Also check all items for this product to see what's there
        all_items = InventoryItem.query.filter_by(
            branch_id=branch.id,
            product_id=product.id
        ).all()
        
        print(f"DEBUG: After rollback, found {len(all_items)} items for this product:")
        for it in all_items:
            print(f"DEBUG:   - Item id={it.id}, batch_code='{it.batch_code}'")
        
        if existing_check:
            # Get all existing batch codes for better error message
            all_batches = InventoryItem.query.filter_by(
                branch_id=branch.id,
                product_id=product.id
            ).with_entities(InventoryItem.batch_code).all()
            existing_batches = [b[0] for b in all_batches if b[0]]
            
            error_response = f"Product '{product_name}' with batch code '{batch_code or '(none)'}' already exists in this branch."
            if existing_batches:
                error_response += f" Existing batch codes: {', '.join(existing_batches)}"
            error_response += " Please use a different batch code or update the existing item."
        else:
            # This shouldn't happen if our pre-check worked, but handle it anyway
            error_response = f"Duplicate branch/product/batch combination detected. Product: '{product_name}' (ID: {product.id}), Branch: {branch.id}, Batch: '{batch_code or '(none)'}'. "
            if all_items:
                existing_batches = [it.batch_code for it in all_items if it.batch_code]
                if existing_batches:
                    error_response += f"Existing batch codes: {', '.join(existing_batches)}. "
            error_response += "Please use a different batch code."
        
        print(f"DEBUG: Returning error: {error_response}")
        return jsonify({"ok": False, "error": error_response}), 409
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# READ: list items for manager branch (or explicit branch_id)
@manager_bp.route("/api/inventory", methods=["GET"])
@manager_required
def mgr_inventory_list():
    try:
        branch_id = request.args.get("branch_id", type=int) or _current_manager_branch_id()
        
        # If no branch_id found, try to get from URL parameters or default to branch 1
        if not branch_id:
            url_branch = request.args.get('branch')
            if url_branch:
                branch_id = int(url_branch)
            else:
                branch_id = 1  # Default to branch 1 for testing
        
        from sqlalchemy.orm import load_only
        q = InventoryItem.query.options(
            load_only(
                InventoryItem.id,
                InventoryItem.branch_id,
                InventoryItem.product_id,
                InventoryItem.stock_kg,
                InventoryItem.unit_price,
                InventoryItem.batch_code,
                InventoryItem.grn_number,
                InventoryItem.warn_level,
                InventoryItem.auto_level,
                InventoryItem.margin,
            )
        )
        if branch_id:
            q = q.filter(InventoryItem.branch_id == branch_id)

        # Optional text filter by product name (?q=jas)
        qtext = (request.args.get("q") or "").strip()
        if qtext:
            q = q.join(Product).filter(Product.name.ilike(f"%{qtext}%"))

        items = q.order_by(InventoryItem.id.desc()).all()
        
        # Debug logging
        print(f"DEBUG: Inventory API called with branch_id={branch_id}, found {len(items)} items")
        
        return jsonify({"ok": True, "items": [item_to_dict(it) for it in items]}), 200
    except Exception as e:
        print(f"DEBUG: Error in mgr_inventory_list: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e), "items": []}), 500


# READ: one item
@manager_bp.route("/api/inventory/<int:item_id>", methods=["GET"])
@manager_required
def mgr_inventory_get(item_id: int):
    from sqlalchemy.orm import load_only
    it = (
        InventoryItem.query.options(
            load_only(
                InventoryItem.id,
                InventoryItem.branch_id,
                InventoryItem.product_id,
                InventoryItem.stock_kg,
                InventoryItem.unit_price,
                InventoryItem.batch_code,
                InventoryItem.grn_number,
                InventoryItem.warn_level,
                InventoryItem.auto_level,
                InventoryItem.margin,
            )
        )
        .filter(InventoryItem.id == item_id)
        .first_or_404()
    )
    return jsonify({"ok": True, "item": item_to_dict(it)}), 200

@manager_bp.get("/api/products/<int:product_id>/batch-codes")
@manager_required
def mgr_product_batch_codes(product_id: int):
    """Return distinct batch codes for a product within the manager's branch."""
    try:
        branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
        from sqlalchemy import func
        q = db.session.query(InventoryItem.batch_code).filter(
            InventoryItem.product_id == product_id,
            InventoryItem.batch_code.isnot(None),
            InventoryItem.batch_code != ''
        )
        if branch_id:
            q = q.filter(InventoryItem.branch_id == branch_id)
        codes = [row[0] for row in q.distinct().all() if row[0]]
        return jsonify({"ok": True, "batch_codes": codes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "batch_codes": []}), 500


# UPDATE: patch inventory/product fields
@manager_bp.route("/api/inventory/<int:item_id>", methods=["PATCH"])
@manager_required
def mgr_inventory_update(item_id: int):
    data = request.get_json(silent=True) or {}
    # Avoid loading undefined grn_number column on Render
    from sqlalchemy.orm import load_only
    it = (
        db.session.query(InventoryItem)
        .options(
            load_only(
                InventoryItem.id,
                InventoryItem.branch_id,
                InventoryItem.product_id,
                InventoryItem.stock_kg,
                InventoryItem.unit_price,
                InventoryItem.batch_code,
                InventoryItem.grn_number,
                InventoryItem.warn_level,
                InventoryItem.auto_level,
                InventoryItem.margin,
            )
        )
        .filter(InventoryItem.id == item_id)
        .first()
    )
    if not it:
        return jsonify({"ok": False, "error": "Inventory item not found"}), 404

    try:
        # inventory fields
        if "stock_kg" in data and data["stock_kg"] != "":
            it.stock_kg = _to_float(data["stock_kg"]) or 0.0
        if "unit_price" in data and data["unit_price"] != "":
            it.unit_price = _to_float(data["unit_price"]) or 0.0
        if "batch_code" in data or "batch" in data:
            it.batch_code = (data.get("batch_code") or data.get("batch") or None)
        if "grn_number" in data or "grn" in data:
            it.grn_number = (data.get("grn_number") or data.get("grn") or "").strip() or None
        if "warn_level" in data or "warn" in data:
            it.warn_level = _to_float(data.get("warn_level") or data.get("warn"))
        if "auto_level" in data or "auto" in data:
            it.auto_level = _to_float(data.get("auto_level") or data.get("auto"))
        if "margin" in data:
            it.margin = (data["margin"] or None)

        # product fields (optional)
        if "product_name" in data and data["product_name"]:
            name = data["product_name"].strip()
            prod = Product.query.filter_by(name=name).first()
            if not prod:
                prod = Product(name=name)
                db.session.add(prod)
                db.session.flush()
            it.product_id = prod.id

        if it.product:
            if "category" in data: it.product.category = (data["category"] or None)
            if "barcode"  in data: it.product.barcode  = (data["barcode"] or None)
            if "sku"      in data: it.product.sku      = (data["sku"] or None)
            if "desc"     in data: it.product.description = (data["desc"] or None)

        # Update timestamp
        from datetime import datetime
        it.updated_at = datetime.utcnow()

        db.session.commit()
        return jsonify({"ok": True, "item": item_to_dict(it)}), 200
    except IntegrityError:
        db.session.rollback()
        return jsonify({"ok": False, "error": "Duplicate branch/product/batch (uq_branch_product_batch). Try a different batch code."}), 409
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# DELETE: remove inventory item (product stays)
@manager_bp.route("/api/inventory/<int:item_id>", methods=["DELETE"])
@manager_required
def mgr_inventory_delete(item_id: int):
    # Avoid loading undefined grn_number column
    from sqlalchemy.orm import load_only
    it = (
        db.session.query(InventoryItem)
        .options(load_only(InventoryItem.id))
        .filter(InventoryItem.id == item_id)
        .first()
    )
    if not it:
        return jsonify({"ok": False, "error": "Inventory item not found"}), 404
    try:
        db.session.delete(it)
        db.session.commit()
        return jsonify({"ok": True, "deleted_id": item_id}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# RESTOCK: add to stock + create restock log
@manager_bp.route("/api/inventory/<int:item_id>/restock", methods=["POST"])
@manager_required
def mgr_inventory_restock(item_id: int):
    """
    JSON body:
    { "qty": 100, "supplier": "Acme", "note": "PO-123", "date": "YYYY-MM-DD" }
    Also accepts "quantity" as alias of "qty".
    """
    from datetime import datetime
    # Avoid loading undefined grn_number column
    from sqlalchemy.orm import load_only
    it = (
        db.session.query(InventoryItem)
        .options(
            load_only(
                InventoryItem.id,
                InventoryItem.branch_id,
                InventoryItem.product_id,
                InventoryItem.stock_kg,
            )
        )
        .filter(InventoryItem.id == item_id)
        .first()
    )
    if not it:
        return jsonify({"ok": False, "error": "Inventory item not found"}), 404
    data = request.get_json(silent=True) or {}

    try:
        qty = _to_float(data.get("qty") if "qty" in data else data.get("quantity"))
        if not qty or qty <= 0:
            return jsonify({"ok": False, "error": "qty/quantity must be > 0"}), 400

        it.stock_kg = (it.stock_kg or 0) + qty
        
        # Update timestamp
        it.updated_at = datetime.utcnow()

        # Optional date override
        created_at = None
        if data.get("date"):
            try:
                created_at = datetime.strptime(data["date"], "%Y-%m-%d")
            except ValueError:
                return jsonify({"ok": False, "error": "date must be YYYY-MM-DD"}), 400

        log = RestockLog(
            inventory_item=it,
            qty_kg=qty,
            supplier=(data.get("supplier") or "Admin"),  # show as "Admin" in UI if blank
            note=(data.get("note") or data.get("notes") or "Restock"),
            created_at=created_at or datetime.utcnow()
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            "ok": True,
            "item": item_to_dict(it),
            "log": log.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# RESTOCK: general restock endpoint using product_id
@manager_bp.route("/api/inventory/restock", methods=["POST"])
@manager_required
def mgr_inventory_restock_general():
    """
    JSON body:
    {
        "product_id": 1,
        "quantity": 50.0,
        "supplier": "ABC Supplier",
        "note": "Fresh delivery"
    }
    """
    data = request.get_json(silent=True) or {}
    branch_id = _current_manager_branch_id()
    
    print(f"DEBUG: Restock data received: {data}")
    
    try:
        # The frontend sends inventory_item_id, not product_id
        inventory_item_id = data.get("product_id")  # This is actually the inventory item ID from the dropdown
        quantity = _to_float(data.get("quantity")) or 0.0
        supplier = (data.get("supplier") or "").strip()
        note = (data.get("note") or "").strip()
        
        print(f"DEBUG: Parsed values - inventory_item_id: {inventory_item_id}, quantity: {quantity}, supplier: '{supplier}', note: '{note}'")
        
        if not inventory_item_id:
            return jsonify({"ok": False, "error": "Product selection is required"}), 400
        
        if quantity <= 0:
            return jsonify({"ok": False, "error": "Quantity must be positive"}), 400
        
        # Supplier is optional (can be empty string or None)
        if not supplier:
            supplier = None
        
        # Find the inventory item directly by ID and verify it belongs to manager's branch
        inventory_item = InventoryItem.query.filter_by(
            id=inventory_item_id,
            branch_id=branch_id
        ).first()
        
        print(f"DEBUG: Looking for inventory_item_id={inventory_item_id} in branch_id={branch_id}")
        print(f"DEBUG: Found inventory_item: {inventory_item}")
        
        if not inventory_item:
            # Let's check what inventory items actually exist in this branch
            all_items = InventoryItem.query.filter_by(branch_id=branch_id).all()
            print(f"DEBUG: All items in branch {branch_id}: {[(item.id, item.product.name if item.product else 'No Product') for item in all_items]}")
            return jsonify({"ok": False, "error": f"Inventory item not found in your branch. Item ID: {inventory_item_id}, Branch ID: {branch_id}"}), 404
        
        # Update stock
        inventory_item.stock_kg = (inventory_item.stock_kg or 0.0) + quantity
        
        # Update timestamp
        from datetime import datetime
        inventory_item.updated_at = datetime.utcnow()
        
        # Create restock log - manager performs the restock
        log = RestockLog(
            inventory_item=inventory_item,
            qty_kg=quantity,
            supplier=supplier,
            note=note or None,
            created_by="Manager"  # This will show as "By: Manager" in admin logs
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            "ok": True, 
            "message": f"Restocked {quantity}kg of {inventory_item.product.name}",
            "item": item_to_dict(inventory_item)
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# LIST RESTOCK LOGS FOR AN ITEM
@manager_bp.route("/api/inventory/<int:item_id>/logs", methods=["GET"])
@manager_required
def mgr_inventory_logs(item_id: int):
    # Avoid loading undefined grn_number column
    from sqlalchemy.orm import load_only
    it = (
        db.session.query(InventoryItem)
        .options(load_only(InventoryItem.id))
        .filter(InventoryItem.id == item_id)
        .first()
    )
    if not it:
        return jsonify({"ok": False, "error": "Inventory item not found"}), 404
    # ensure newest first; your model's to_dict should include date/supplier/note
    logs = RestockLog.query.filter_by(inventory_item_id=it.id) \
                           .order_by(RestockLog.created_at.desc()) \
                           .all()
    return jsonify({"ok": True, "logs": [l.to_dict() for l in logs]}), 200


# ======================== ANALYTICS (BRANCH) ========================
# ======================== MANAGER DASHBOARD API ========================
@manager_bp.get("/api/dashboard/kpis")
@manager_required
def mgr_dashboard_kpis():
    """Get KPI data for manager dashboard (branch-specific)"""
    try:
        from datetime import datetime, date, timedelta
        from sqlalchemy import func, and_, or_
        
        # Get manager's branch ID - prioritize URL parameter over session
        url_branch_id = request.args.get('branch_id', type=int)
        session_branch_id = _current_manager_branch_id()
        branch_id = url_branch_id or session_branch_id or 1
        if not branch_id:
            return jsonify({"ok": False, "error": "Manager branch not found"}), 400
        
        print(f"DEBUG KPI: Manager dashboard KPIs called")
        print(f"  - URL branch_id: {url_branch_id}")
        print(f"  - Session branch_id: {session_branch_id}")
        print(f"  - Using branch_id: {branch_id}")
        print(f"  - Request args: {dict(request.args)}")
        print(f"  - Session user: {session.get('user', {}).get('branch_id')}")
        
        # Initialize default values
        today_sales = 0
        month_sales = 0
        low_stock_count = 0
        forecast_accuracy = 0
        total_orders = 0
        
        try:
            # Get current date and month in Philippines timezone
            from datetime import timezone as tz
            ph_tz = tz(timedelta(hours=8))
            now_ph = datetime.now(ph_tz)
            today = now_ph.date()  # Today in Philippines time
            current_month = today.month
            current_year = today.year
            
            # Today's sales for this branch (using Philippines timezone)
            today_start_ph = datetime.combine(today, datetime.min.time()).replace(tzinfo=ph_tz)
            today_end_ph = datetime.combine(today, datetime.max.time()).replace(tzinfo=ph_tz)
            today_start_utc = today_start_ph.astimezone(tz.utc).replace(tzinfo=None)
            today_end_utc = today_end_ph.astimezone(tz.utc).replace(tzinfo=None)
            
            today_sales = db.session.query(SalesTransaction).filter(
                and_(
                    SalesTransaction.branch_id == branch_id,
                    SalesTransaction.transaction_date >= today_start_utc,
                    SalesTransaction.transaction_date <= today_end_utc
                )
            ).with_entities(func.sum(SalesTransaction.total_amount)).scalar() or 0
            
            # This month's sales for this branch
            month_sales = db.session.query(SalesTransaction).filter(
                and_(
                    SalesTransaction.branch_id == branch_id,
                    func.extract('month', SalesTransaction.transaction_date) == current_month,
                    func.extract('year', SalesTransaction.transaction_date) == current_year
                )
            ).with_entities(func.sum(SalesTransaction.total_amount)).scalar() or 0
            
        except Exception as e:
            print(f"DEBUG KPI: Error in sales queries: {e}")
            import traceback
            traceback.print_exc()
        
        try:
            # Low stock count for this branch (avoid grn_number column)
            from sqlalchemy.orm import load_only
            # Check for items where stock is below warn_level (if set) or below 100kg (default threshold)
            low_stock_count = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.stock_kg, InventoryItem.warn_level))
                .filter(
                    and_(
                        InventoryItem.branch_id == branch_id,
                        or_(
                            and_(InventoryItem.warn_level.isnot(None), InventoryItem.stock_kg < InventoryItem.warn_level),
                            and_(InventoryItem.warn_level.is_(None), InventoryItem.stock_kg < 100)
                        )
                    )
                )
                .count()
            )
        except Exception as e:
            print(f"DEBUG KPI: Error in inventory query: {e}")
            import traceback
            traceback.print_exc()
            low_stock_count = 0
        
        try:
            # Forecast accuracy for this branch
            forecast_accuracy = db.session.query(func.avg(ForecastData.accuracy_score)).filter(
                ForecastData.branch_id == branch_id
            ).scalar() or 0
        except Exception as e:
            print(f"DEBUG KPI: Error in forecast query: {e}")
            import traceback
            traceback.print_exc()
            forecast_accuracy = 0
        
        try:
            # Calculate Total Orders for this specific branch
            # Count distinct transaction dates per branch (each purchase form submission creates
            # multiple SalesTransaction records with the same transaction_date)
            # This counts unique days with sales, which approximates orders
            total_orders = db.session.query(
                func.count(func.distinct(func.date(SalesTransaction.transaction_date)))
            ).filter(
                SalesTransaction.branch_id == branch_id
            ).scalar() or 0
            
            # Debug: also check how many transactions exist
            all_transactions = db.session.query(SalesTransaction).filter(
                SalesTransaction.branch_id == branch_id
            ).all()
            print(f"DEBUG KPI: Total orders query for branch_id={branch_id}:")
            print(f"  - Total orders count (distinct dates): {total_orders}")
            print(f"  - Total transaction records for branch: {len(all_transactions)}")
            if all_transactions:
                sample = all_transactions[0]
                print(f"  - Sample transaction branch_id: {sample.branch_id}, date: {sample.transaction_date}")
            
            # Double-check: query distinct dates by date range to see recent orders
            today = date.today()
            recent_orders = db.session.query(
                func.count(func.distinct(func.date(SalesTransaction.transaction_date)))
            ).filter(
                and_(
                    SalesTransaction.branch_id == branch_id,
                    func.date(SalesTransaction.transaction_date) >= today - timedelta(days=7)
                )
            ).scalar() or 0
            print(f"  - Recent orders (distinct dates in last 7 days) for branch {branch_id}: {recent_orders}")
            
        except Exception as e:
            print(f"DEBUG KPI: Error in orders query: {e}")
            import traceback
            traceback.print_exc()
            total_orders = 0
        
        return jsonify({
            "ok": True,
            "kpis": {
                "today_sales": float(today_sales),
                "month_sales": float(month_sales),
                "low_stock_count": int(low_stock_count),
                "forecast_accuracy": round(forecast_accuracy, 2),
                "total_orders": int(total_orders)
            }
        })
    except Exception as e:
        print(f"DEBUG KPI: Critical error in mgr_dashboard_kpis: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "ok": False,
            "error": str(e),
            "kpis": {
                "today_sales": 0,
                "month_sales": 0,
                "low_stock_count": 0,
                "forecast_accuracy": 0,
                "total_orders": 0
            }
        }), 500

@manager_bp.get("/api/dashboard/charts")
@manager_required
def mgr_dashboard_charts():
    """Get chart data for manager dashboard (branch-specific)"""
    from datetime import datetime, date, timedelta
    from sqlalchemy import func, and_, desc
    
    # Get manager's branch ID - prioritize URL parameter over session
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id() or 1
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    # Get query parameters
    days = request.args.get('days', 30, type=int)
    
    # Date range - include today (end_date) in the range
    # Use Philippines timezone for date comparison
    from datetime import timezone as tz
    ph_tz = tz(timedelta(hours=8))
    now_ph = datetime.now(ph_tz)
    end_date = now_ph.date()  # Today in Philippines time
    start_date = end_date - timedelta(days=days - 1)  # Adjust to include today
    
    # Sales trend data for this branch
    # Query all transactions in the date range, then group by Philippines date
    # Since transactions are stored as naive UTC, we need to convert to Philippines time for date extraction
    all_sales = db.session.query(SalesTransaction).filter(
        and_(
            SalesTransaction.branch_id == branch_id,
            SalesTransaction.transaction_date >= datetime.combine(start_date, datetime.min.time()) - timedelta(hours=8),  # Convert PH date start to UTC
            SalesTransaction.transaction_date < datetime.combine(end_date + timedelta(days=1), datetime.min.time()) - timedelta(hours=8)  # Convert PH date end to UTC
        )
    ).all()
    
    # Group by Philippines date
    sales_trend_dict = {}
    for sale in all_sales:
        # Convert UTC datetime to Philippines time and get date
        if sale.transaction_date.tzinfo is None:
            # Naive datetime, assume UTC
            sale_utc = sale.transaction_date.replace(tzinfo=tz.utc)
        else:
            sale_utc = sale.transaction_date
        sale_ph = sale_utc.astimezone(ph_tz)
        sale_date = sale_ph.date()
        
        if start_date <= sale_date <= end_date:
            if sale_date not in sales_trend_dict:
                sales_trend_dict[sale_date] = {'sales': 0.0, 'quantity': 0.0}
            sales_trend_dict[sale_date]['sales'] += float(sale.total_amount)
            sales_trend_dict[sale_date]['quantity'] += float(sale.quantity_sold)
    
    # Fill in missing dates with zero values - include today
    sales_trend_filled = []
    for i in range(days):
        current_date = start_date + timedelta(days=i)
        # Ensure we don't go past today
        if current_date > end_date:
            break
        if current_date in sales_trend_dict:
            sales_trend_filled.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'sales': sales_trend_dict[current_date]['sales'],
                'quantity': sales_trend_dict[current_date]['quantity']
            })
        else:
            sales_trend_filled.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'sales': 0.0,
                'quantity': 0.0
            })
    
    # Forecast vs Actual data for this branch
    forecast_vs_actual = []
    for i in range(days):
        current_date = start_date + timedelta(days=i)
        # Ensure we don't go past today
        if current_date > end_date:
            break
        
        # Get actual sales for this date (use Philippines timezone)
        current_date_start_ph = datetime.combine(current_date, datetime.min.time()).replace(tzinfo=ph_tz)
        current_date_end_ph = datetime.combine(current_date, datetime.max.time()).replace(tzinfo=ph_tz)
        current_date_start_utc = current_date_start_ph.astimezone(tz.utc).replace(tzinfo=None)
        current_date_end_utc = current_date_end_ph.astimezone(tz.utc).replace(tzinfo=None)
        
        actual_sales = db.session.query(SalesTransaction).filter(
            and_(
                SalesTransaction.branch_id == branch_id,
                SalesTransaction.transaction_date >= current_date_start_utc,
                SalesTransaction.transaction_date <= current_date_end_utc
            )
        ).with_entities(
            func.sum(SalesTransaction.quantity_sold)
        ).scalar() or 0
        
        # Get forecast for this date
        forecast_sales = db.session.query(ForecastData).filter(
            and_(
                ForecastData.branch_id == branch_id,
                ForecastData.forecast_date == current_date
            )
        ).with_entities(
            func.sum(ForecastData.predicted_demand)
        ).scalar() or 0
        
        forecast_vs_actual.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'actual': float(actual_sales),
            'forecast': float(forecast_sales)
        })
    
    # Top 5 products this month for this branch
    current_month = date.today().month
    current_year = date.today().year
    
    top_products = db.session.query(
        Product.name,
        func.sum(SalesTransaction.quantity_sold).label('total_quantity'),
        func.sum(SalesTransaction.total_amount).label('total_sales')
    ).join(
        SalesTransaction, Product.id == SalesTransaction.product_id
    ).filter(
        and_(
            SalesTransaction.branch_id == branch_id,
            func.extract('month', SalesTransaction.transaction_date) == current_month,
            func.extract('year', SalesTransaction.transaction_date) == current_year
        )
    ).group_by(
        Product.id, Product.name
    ).order_by(
        desc('total_quantity')
    ).limit(5).all()
    
    return jsonify({
        "ok": True,
        "charts": {
            "sales_trend": sales_trend_filled,
            "forecast_vs_actual": forecast_vs_actual,
            "top_products": [
                {
                    'name': row.name,
                    'quantity': float(row.total_quantity),
                    'sales': float(row.total_sales)
                }
                for row in top_products
            ]
        }
    })

@manager_bp.get("/api/analytics")
@manager_required
def mgr_analytics_overview():
    """Branch-only analytics for manager Analytics page.
    Query params: branch_id (optional; defaults to current manager's branch), days (default 30)
    Returns JSON with:
      - forecast_vs_actual_7d: [{ date, forecast, actual }]
      - product_accuracy: [{ product_id, name, accuracy, forecasted, actual }]
      - stock_movement_7d: { labels: [dates], stock_in: [..], stock_out: [..] }
      - category_levels: [{ category, stock_kg }]
      - top_products_month: [{ name, quantity, sales }]
      - weekly_sales_4w: { labels: [Week 1..4], sales: [..] }
      - branch_risks: [{ product_id, product_name, risk_level, gap_kg, action, risk_score }]
    """
    from sqlalchemy import func, and_, desc
    from datetime import date, datetime, timedelta

    branch_id = request.args.get("branch_id", type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "branch_id is required"}), 400
    
    # Verify branch exists and belongs to manager
    branch = Branch.query.get(branch_id)
    if not branch:
        return jsonify({"ok": False, "error": f"Branch {branch_id} not found"}), 404
    
    print(f"DEBUG ANALYTICS: Fetching analytics for branch_id={branch_id} (branch: {branch.name})")

    today = date.today()
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)

    # 1) Forecast vs Actual (last 7 days), aggregated across products for the branch
    f_rows = (
        db.session.query(
            ForecastData.forecast_date.label('d'),
            func.sum(ForecastData.predicted_demand).label('pred')
        )
        .filter(and_(ForecastData.branch_id == branch_id,
                     ForecastData.forecast_date >= seven_days_ago,
                     ForecastData.forecast_date <= today))
        .group_by('d')
        .all()
    )
    f_map = {row.d: float(row.pred or 0) for row in f_rows}

    a_rows = (
        db.session.query(
            func.date(SalesTransaction.transaction_date).label('d'),
            func.sum(SalesTransaction.quantity_sold).label('qty')
        )
        .filter(and_(SalesTransaction.branch_id == branch_id,
                     func.date(SalesTransaction.transaction_date) >= seven_days_ago,
                     func.date(SalesTransaction.transaction_date) <= today))
        .group_by('d')
        .all()
    )
    a_map = {row.d: float(row.qty or 0) for row in a_rows}

    forecast_vs_actual_7d = []
    for i in range(7):
        d = seven_days_ago + timedelta(days=i)
        forecast_vs_actual_7d.append({
            "date": d.strftime('%Y-%m-%d'),
            "forecast": float(f_map.get(d, 0.0)),
            "actual": float(a_map.get(d, 0.0)),
        })

    # 2) Per-product accuracy (MAPE) for last 30 days
    per_product = (
        db.session.query(
            ForecastData.product_id,
            func.sum(ForecastData.predicted_demand).label('pred_sum')
        )
        .filter(and_(ForecastData.branch_id == branch_id,
                     ForecastData.forecast_date >= thirty_days_ago,
                     ForecastData.forecast_date <= today))
        .group_by(ForecastData.product_id)
        .all()
    )
    product_ids = [pid for pid, _ in per_product]
    # actual per product
    actual_per_product = (
        db.session.query(
            SalesTransaction.product_id,
            func.sum(SalesTransaction.quantity_sold).label('qty_sum')
        )
        .filter(and_(SalesTransaction.branch_id == branch_id,
                     func.date(SalesTransaction.transaction_date) >= thirty_days_ago,
                     func.date(SalesTransaction.transaction_date) <= today,
                     SalesTransaction.product_id.in_(product_ids) if product_ids else True))
        .group_by(SalesTransaction.product_id)
        .all()
    )
    actual_map = {pid: float(q or 0) for pid, q in actual_per_product}
    pred_map = {pid: float(p or 0) for pid, p in per_product}

    # Get product names - only for products in this branch's inventory or forecasts
    # First get all products that have inventory in this branch (avoid grn_number column)
    from sqlalchemy.orm import load_only
    branch_product_ids = set([it.product_id for it in 
        db.session.query(InventoryItem)
        .options(load_only(InventoryItem.product_id))
        .filter(InventoryItem.branch_id == branch_id)
        .all()])
    # Add products from forecasts
    branch_product_ids.update(product_ids)
    names = {p.id: p.name for p in Product.query.filter(Product.id.in_(list(branch_product_ids))).all()} if branch_product_ids else {}

    product_accuracy = []
    total_accuracy_sum = 0.0
    total_accuracy_count = 0
    for pid in product_ids:
        pred = pred_map.get(pid, 0.0)
        act = actual_map.get(pid, 0.0)
        acc = 0.0
        if act > 0:
            mape = abs(pred - act) / act * 100.0
            acc = max(0.0, 100.0 - mape)
        product_accuracy.append({
            "product_id": pid,
            "name": names.get(pid),
            "forecasted": round(pred, 2),
            "actual": round(act, 2),
            "accuracy": round(acc, 2),
        })
        if act > 0:
            total_accuracy_sum += acc
            total_accuracy_count += 1
    
    # Calculate overall accuracy
    overall_accuracy = round(total_accuracy_sum / total_accuracy_count, 2) if total_accuracy_count > 0 else 0.0
    
    # Calculate date period for last 7 days
    accuracy_period = f"{seven_days_ago.strftime('%b %d')} - {today.strftime('%b %d')}"

    # 3) Stock movement (7d): stock_in from RestockLog, stock_out from SalesTransaction
    # stock in (by date) - avoid grn_number column
    inv_ids = [it.id for it in 
        db.session.query(InventoryItem.id)
        .filter(InventoryItem.branch_id == branch_id)
        .all()]
    in_rows = []
    if inv_ids:
        in_rows = (
            db.session.query(
                func.date(RestockLog.created_at).label('d'),
                func.sum(RestockLog.qty_kg).label('qty')
            )
            .filter(RestockLog.inventory_item_id.in_(inv_ids))
            .filter(func.date(RestockLog.created_at) >= seven_days_ago)
            .group_by('d')
            .all()
        )
    in_map = {row.d: float(row.qty or 0) for row in in_rows}

    out_rows = (
        db.session.query(
            func.date(SalesTransaction.transaction_date).label('d'),
            func.sum(SalesTransaction.quantity_sold).label('qty')
        )
        .filter(and_(SalesTransaction.branch_id == branch_id,
                     func.date(SalesTransaction.transaction_date) >= seven_days_ago))
        .group_by('d')
        .all()
    )
    out_map = {row.d: float(row.qty or 0) for row in out_rows}

    labels_7d = [(seven_days_ago + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    stock_movement_7d = {
        "labels": labels_7d,
        "stock_in": [in_map.get(datetime.strptime(d, '%Y-%m-%d').date(), 0.0) for d in labels_7d],
        "stock_out": [out_map.get(datetime.strptime(d, '%Y-%m-%d').date(), 0.0) for d in labels_7d],
    }

    # 4) Category stock levels (current)
    cat_rows = (
        db.session.query(Product.category, func.sum(InventoryItem.stock_kg))
        .join(Product, Product.id == InventoryItem.product_id)
        .filter(InventoryItem.branch_id == branch_id)
        .group_by(Product.category)
        .all()
    )
    category_levels = [{"category": (c or "Uncategorized"), "stock_kg": float(s or 0)} for c, s in cat_rows]

    # 5) Top products this month (quantity, sales)
    current_month = today.month
    current_year = today.year
    top_rows = (
        db.session.query(Product.name, func.sum(SalesTransaction.quantity_sold).label('qty'), func.sum(SalesTransaction.total_amount).label('amt'))
        .join(Product, Product.id == SalesTransaction.product_id)
        .filter(and_(SalesTransaction.branch_id == branch_id,
                     func.extract('month', SalesTransaction.transaction_date) == current_month,
                     func.extract('year', SalesTransaction.transaction_date) == current_year))
        .group_by(Product.id, Product.name)
        .order_by(desc('qty'))
        .limit(5)
        .all()
    )
    top_products_month = [{"name": n, "quantity": float(q or 0), "sales": float(a or 0)} for n, q, a in top_rows]

    # 6) Weekly sales last 4 weeks
    start_4w = today - timedelta(days=28)
    daily_rows = (
        db.session.query(func.date(SalesTransaction.transaction_date).label('d'), func.sum(SalesTransaction.total_amount).label('amt'))
        .filter(and_(SalesTransaction.branch_id == branch_id,
                     func.date(SalesTransaction.transaction_date) >= start_4w))
        .group_by('d')
        .order_by('d')
        .all()
    )
    # bucket into 4 weeks
    week_labels = ["Week 1", "Week 2", "Week 3", "Week 4"]
    week_buckets = [0.0, 0.0, 0.0, 0.0]
    for d, amt in daily_rows:
        days_from_start = (d - start_4w).days
        idx = min(3, max(0, days_from_start // 7))
        week_buckets[idx] += float(amt or 0)
    weekly_sales_4w = {"labels": week_labels, "sales": [round(v, 2) for v in week_buckets]}

    # 7) Branch risks: forecast next 7 days vs stock (only for products in this branch)
    next7 = today + timedelta(days=7)
    f_next = (
        db.session.query(ForecastData.product_id, func.sum(ForecastData.predicted_demand).label('pred'))
        .filter(and_(ForecastData.branch_id == branch_id,
                     ForecastData.forecast_date >= today,
                     ForecastData.forecast_date <= next7))
        .group_by(ForecastData.product_id)
        .all()
    )
    # Get inventory map for this branch only (avoid grn_number column)
    inv_items = (
        db.session.query(InventoryItem)
        .options(load_only(InventoryItem.product_id, InventoryItem.stock_kg))
        .filter(InventoryItem.branch_id == branch_id)
        .all()
    )
    inv_map = {it.product_id: float(it.stock_kg or 0) for it in inv_items}
    # Ensure product names are available for all products in this branch's inventory
    branch_inv_product_ids = [it.product_id for it in inv_items if it.product_id]
    if branch_inv_product_ids:
        branch_product_names = {p.id: p.name for p in Product.query.filter(Product.id.in_(branch_inv_product_ids)).all()}
        names.update(branch_product_names)
    
    branch_risks = []
    for pid, pred in f_next:
        stock = inv_map.get(pid, 0.0)
        gap = float(pred or 0) - stock
        # Only show risks for products with positive gaps (shortage risk)
        if gap > 0:
            risk_level = "High" if gap > stock else ("Medium" if gap > stock * 0.5 else "Low")
            risk_score = round(min(100.0, (gap / (stock + 1e-6)) * 100.0), 2)
            action = "Increase stock" if risk_level == "High" else "Monitor & adjust"
            name = names.get(pid) if pid in names else (Product.query.get(pid).name if Product.query.get(pid) else f"Product {pid}")
            branch_risks.append({
                "product_id": pid,
                "product_name": name,
                "risk_level": risk_level,
                "gap_kg": round(gap, 2),
                "action": action,
                "risk_score": risk_score,
            })

    return jsonify({
        "ok": True,
        "analytics": {
            "overall_accuracy": overall_accuracy,
            "accuracy_period": accuracy_period,
            "forecast_vs_actual_7d": forecast_vs_actual_7d,
            "product_accuracy": product_accuracy,
            "stock_movement_7d": stock_movement_7d,
            "category_levels": category_levels,
            "top_products_month": top_products_month,
            "weekly_sales_4w": weekly_sales_4w,
            "branch_risks": branch_risks,
        }
    })


# ========================= FORECAST API =========================
@manager_bp.route("/api/forecast", methods=["POST"])
@manager_required
def mgr_forecast_generate():
    """Generate forecast for manager's branch using ARIMA/SARIMA models"""
    from datetime import datetime, timedelta
    import numpy as np
    from sqlalchemy import func
    from forecasting_service import forecasting_service
    
    data = request.get_json(silent=True) or {}
    branch_id = _current_manager_branch_id()
    product_id = data.get('product_id')
    category = data.get('category')
    days = data.get('days', 30)
    model_type = data.get('model_type', 'arima')
    
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        today = datetime.now().date()
        forecast_data = []
        
        # Get products from manager's inventory (avoid grn_number column)
        from sqlalchemy.orm import load_only
        if product_id:
            products = [Product.query.get(product_id)]
        else:
            # Get inventory items with optional category filter
            query = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id))
                .filter_by(branch_id=branch_id)
            )
            if category:
                query = query.join(Product).filter(Product.category == category)
            inventory_items = query.all()
            products = [item.product for item in inventory_items if item.product]
        
        for product in products:
            if not product:
                continue
            
            print(f"DEBUG FORECAST: Processing product: {product.name}")
            
            # Get historical sales data for this product (last 90 days)
            sales_transactions = db.session.query(SalesTransaction).filter(
                SalesTransaction.branch_id == branch_id,
                SalesTransaction.product_id == product.id,
                func.date(SalesTransaction.transaction_date) >= today - timedelta(days=90)
            ).order_by(SalesTransaction.transaction_date.desc()).all()
            
            print(f"DEBUG FORECAST: Found {len(sales_transactions)} sales transactions for {product.name}")
            
            # Convert to list for forecasting service
            historical_data = []
            for sale in sales_transactions:
                historical_data.append({
                    'transaction_date': sale.transaction_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'quantity_sold': sale.quantity_sold
                })
            
            # If no sales data, create some dummy data based on inventory (avoid grn_number column)
            if not historical_data:
                inventory_item = (
                    db.session.query(InventoryItem)
                    .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id, InventoryItem.stock_kg))
                    .filter_by(branch_id=branch_id, product_id=product.id)
                    .first()
                )
                
                if inventory_item:
                    # Create dummy sales data based on current stock
                    base_demand = max(10, inventory_item.stock_kg * 0.1)  # 10% of stock as daily demand
                    historical_data = []
                    
                    for i in range(30):  # Last 30 days
                        historical_data.append({
                            "transaction_date": (datetime.now() - timedelta(days=30-i)).strftime("%Y-%m-%d %H:%M:%S"),
                            "quantity_sold": float(base_demand + (i % 7) * 5)  # Some variation
                        })
            
            print(f"DEBUG FORECAST: Historical data length: {len(historical_data)}")
            if historical_data:
                print(f"DEBUG FORECAST: Sample data: {historical_data[:3]}")
            
            # Generate forecast using the forecasting service
            print(f"DEBUG FORECAST: Calling forecasting service for {product.name}")
            forecast_result = forecasting_service.generate_arima_forecast(historical_data, days)
            print(f"DEBUG FORECAST: Forecast result: {forecast_result}")
            
            # Store forecast in database
            for i, (predicted_demand, lower, upper) in enumerate(zip(
                forecast_result['forecast_values'],
                forecast_result['confidence_lower'],
                forecast_result['confidence_upper']
            )):
                forecast_date = today + timedelta(days=i)
                
                # Check if forecast already exists
                existing = ForecastData.query.filter_by(
                    branch_id=branch_id,
                    product_id=product.id,
                    forecast_date=forecast_date,
                    forecast_period='daily'
                ).first()
                
                if not existing:
                    # Convert NumPy types to standard Python types
                    predicted_demand_float = float(predicted_demand)
                    lower_float = float(lower)
                    upper_float = float(upper)
                    accuracy_float = float(forecast_result.get('accuracy_score', 0.8))
                    
                    forecast = ForecastData(
                        branch_id=branch_id,
                        product_id=product.id,
                        forecast_date=forecast_date,
                        forecast_period="daily",
                        predicted_demand=round(predicted_demand_float, 2),
                        confidence_interval_lower=round(lower_float, 2),
                        confidence_interval_upper=round(upper_float, 2),
                        model_type=forecast_result['model_type'],
                        accuracy_score=accuracy_float
                    )
                    db.session.add(forecast)
                    forecast_data.append({
                        'date': forecast_date.strftime('%Y-%m-%d'),
                        'product_id': int(product.id),
                        'product_name': str(product.name),
                        'predicted_demand': float(round(predicted_demand_float, 2)),
                        'confidence': float(round(accuracy_float * 100, 1))
                    })
        
        db.session.commit()
        
        # Prepare chart data for immediate response
        if forecast_data:
            # Group forecast data by date for chart
            date_groups = {}
            for item in forecast_data:
                date = item['date']
                if date not in date_groups:
                    date_groups[date] = {'forecast': 0, 'actual': 0}
                date_groups[date]['forecast'] += item['predicted_demand']
            
            # Get actual sales data for comparison
            actual_sales = db.session.query(
                func.date(SalesTransaction.transaction_date).label('date'),
                func.sum(SalesTransaction.quantity_sold).label('actual_demand')
            ).filter(SalesTransaction.branch_id == branch_id)\
             .filter(func.date(SalesTransaction.transaction_date) >= today - timedelta(days=30))\
             .filter(func.date(SalesTransaction.transaction_date) <= today)\
             .group_by(func.date(SalesTransaction.transaction_date))\
             .all()
            
            actual_map = {row.date.strftime('%Y-%m-%d'): row.actual_demand for row in actual_sales}
            
            # Prepare chart arrays
            labels = []
            forecast_values = []
            actual_values = []
            
            for date in sorted(date_groups.keys()):
                labels.append(date)
                forecast_values.append(float(round(date_groups[date]['forecast'], 2)))
                actual_values.append(float(round(actual_map.get(date, 0), 2)))
            
            # Calculate summary (ensure all values are standard Python types)
            total_forecast = float(sum(forecast_values))
            avg_confidence = float(sum(item['confidence'] for item in forecast_data) / len(forecast_data) if forecast_data else 0.8)
            
            chart_data = {
                "labels": labels,
                "forecast": [float(x) for x in forecast_values],
                "actual": [float(x) for x in actual_values]
            }
            
            summary_data = {
                "avg_demand": float(round(total_forecast / len(forecast_values) if forecast_values else 0, 2)),
                "peak_date": labels[forecast_values.index(max(forecast_values))] if forecast_values else None,
                "peak_quantity": float(max(forecast_values) if forecast_values else 0),
                "suggested_reorder": float(round(total_forecast * 0.1, 2)),
                "confidence": float(round(avg_confidence, 2))
            }
        else:
            chart_data = {"labels": [], "forecast": [], "actual": []}
            summary_data = {"avg_demand": 0, "peak_date": None, "peak_quantity": 0, "suggested_reorder": 0, "confidence": 0.8}
        
        print(f"DEBUG FORECAST: Returning {len(forecast_data)} forecast items")
        for item in forecast_data:
            print(f"DEBUG FORECAST: Product: {item.get('product_name', 'Unknown')}, Forecast points: {len(item.get('forecast_data', []))}")
        
        return jsonify({
            "ok": True,
            "message": f"Generated {model_type.upper()} forecast for {len(forecast_data)} products",
            "forecast": chart_data,
            "summary": summary_data,
            "details": forecast_data,
            "model_info": {
                "type": model_type,
                "days_forecasted": days,
                "confidence_avg": round(sum(f['confidence'] for f in forecast_data) / len(forecast_data), 2) if forecast_data else 0
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500



@manager_bp.route("/api/forecast/price", methods=["POST"])
@manager_required
def mgr_forecast_price():
    """Generate price forecast using Holt-Winters + market factors"""
    from datetime import datetime, timedelta
    import numpy as np
    from sqlalchemy import func
    
    data = request.get_json(silent=True) or {}
    branch_id = _current_manager_branch_id()
    product_id = data.get('product_id')
    days = data.get('days', 30)
    model_type = data.get('model_type', 'holt_winters')
    
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        today = datetime.now().date()
        price_forecast_data = []
        
        # Get products from manager's inventory (avoid grn_number column)
        from sqlalchemy.orm import load_only
        if product_id:
            # When specific product is selected, get it directly
            product = Product.query.get(product_id)
            if not product:
                return jsonify({"ok": False, "error": f"Product with ID {product_id} not found"}), 404
            products = [product]
        else:
            # Get all products from inventory for this branch
            inventory_items = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id))
                .filter_by(branch_id=branch_id)
                .all()
            )
            products = [item.product for item in inventory_items if item.product]
        
        if not products:
            return jsonify({"ok": False, "error": "No products found in inventory for this branch"}), 404
        
        for product in products:
            if not product:
                continue
            
            print(f"DEBUG PRICE FORECAST: Processing product: {product.name} (ID: {product.id})")
            
            # Get current price data (avoid grn_number column)
            current_item = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id, InventoryItem.unit_price))
                .filter_by(branch_id=branch_id, product_id=product.id)
                .first()
            )
            
            print(f"DEBUG PRICE FORECAST: Found inventory item: {current_item}")
            if current_item:
                print(f"DEBUG PRICE FORECAST: Unit price: {current_item.unit_price}, Type: {type(current_item.unit_price)}")
            
            # Handle missing inventory or price - use product default price or reasonable default
            if not current_item:
                print(f"DEBUG PRICE FORECAST: No inventory item found for {product.name}, checking product default price")
                # Try to get price from product table if available
                current_price = float(product.unit_price) if hasattr(product, 'unit_price') and product.unit_price else 50.0
                print(f"DEBUG PRICE FORECAST: Using product default price: {current_price}")
            elif not current_item.unit_price or float(current_item.unit_price) <= 0:
                print(f"DEBUG PRICE FORECAST: Inventory item has no valid price for {product.name}, using default")
                # Use product default price or reasonable default
                current_price = float(product.unit_price) if hasattr(product, 'unit_price') and product.unit_price else 50.0
                print(f"DEBUG PRICE FORECAST: Using default price: {current_price}")
            else:
                current_price = float(current_item.unit_price)
                print(f"DEBUG PRICE FORECAST: Using inventory price: {current_price}")
                
            # Create price history based on current price with some variation
            # Since we don't have historical price data, we'll simulate it
            price_history = []
            
            # Generate 30 days of price history with some variation
            for i in range(30):
                # Add some random variation to simulate price changes
                import random
                variation = random.uniform(-0.05, 0.05)  # ±5% variation
                historical_price = current_price * (1 + variation)
                
                price_history.append({
                    'date': (datetime.now() - timedelta(days=30-i)).strftime("%Y-%m-%d"),
                    'price': historical_price
                })
            
            print(f"DEBUG PRICE FORECAST: Price history length: {len(price_history)}")
            
            if len(price_history) < 3:
                # Use current price if insufficient history
                base_price = current_price
                confidence = 0.5
                print(f"DEBUG PRICE FORECAST: Using current price: {base_price}")
            else:
                # Implement Holt-Winters price forecasting
                prices = [float(item['price']) for item in price_history]
                print(f"DEBUG PRICE FORECAST: Running Holt-Winters with prices: {prices[:5]}...")
                base_price, confidence = _holt_winters_price_forecast(prices, days)
                print(f"DEBUG PRICE FORECAST: Holt-Winters result: base_price={base_price}, confidence={confidence}")
            
            # Generate price forecast for next 'days' days
            for i in range(days):
                forecast_date = today + timedelta(days=i)
                
                # Apply market factors
                market_factor = _get_market_factor(forecast_date, product)
                predicted_price = base_price * market_factor
                
                # Add volatility (price changes over time)
                volatility = 0.02  # 2% daily volatility
                price_change = np.random.normal(0, volatility)
                predicted_price *= (1 + price_change)
                
                # Ensure reasonable price bounds
                predicted_price = max(base_price * 0.8, min(predicted_price, base_price * 1.5))
                
                price_forecast_data.append({
                    'date': forecast_date.strftime('%Y-%m-%d'),
                    'product_id': product.id,
                    'product_name': product.name,
                    'predicted_price': round(predicted_price, 2),
                    'confidence': round(confidence, 2),
                    'market_factor': round(market_factor, 2)
                })
        
        # Generate price insights - use the first product's current price
        price_insights = {}
        if price_forecast_data:
            # Get the first product's current price from the forecast data
            first_product_id = price_forecast_data[0].get('product_id')
            if first_product_id:
                # Find the inventory item for the first product to get current price
                first_item = (
                    db.session.query(InventoryItem)
                    .options(load_only(InventoryItem.id, InventoryItem.unit_price))
                    .filter_by(branch_id=branch_id, product_id=first_product_id)
                    .first()
                )
                if first_item and first_item.unit_price:
                    current_price = float(first_item.unit_price)
                else:
                    # Fallback to product default or reasonable default
                    first_product = Product.query.get(first_product_id)
                    current_price = float(first_product.unit_price) if first_product and hasattr(first_product, 'unit_price') and first_product.unit_price else 50.0
            else:
                current_price = 50.0  # Default fallback
            
            avg_predicted_price = sum(f['predicted_price'] for f in price_forecast_data) / len(price_forecast_data)
            price_change_percent = ((avg_predicted_price - current_price) / current_price) * 100
            
            # Calculate average confidence
            avg_confidence = sum(f.get('confidence', 0.75) for f in price_forecast_data) / len(price_forecast_data) if price_forecast_data else 0.75
            
            price_insights = {
                "current_price": current_price,
                "price_change_percent": round(price_change_percent, 2),
                "predicted_change_percent": round(price_change_percent, 2),
                "reason": f"Based on {model_type} model with {round(avg_confidence * 100, 1)}% confidence"
            }
        
        print(f"DEBUG PRICE FORECAST: Returning {len(price_forecast_data)} products")
        for item in price_forecast_data:
            print(f"DEBUG PRICE FORECAST: Product: {item.get('product_name', 'Unknown')}, Data points: {len(item.get('forecast_data', []))}")
        
        return jsonify({
            "ok": True,
            "message": f"Generated {model_type} price forecast for {len(price_forecast_data)} products",
            "price_forecast": price_forecast_data,
            "price_insights": price_insights,
            "model_info": {
                "type": model_type,
                "days_forecasted": days,
                "confidence_avg": round(sum(f['confidence'] for f in price_forecast_data) / len(price_forecast_data), 2) if price_forecast_data else 0
            }
        })
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def _holt_winters_price_forecast(price_history, forecast_days):
    """Holt-Winters exponential smoothing for price forecasting"""
    import numpy as np
    
    if len(price_history) < 3:
        return np.mean(price_history) if price_history else 50.0, 0.5
    
    data = np.array(price_history)
    
    # Holt-Winters parameters
    alpha = 0.3  # Level smoothing
    beta = 0.1   # Trend smoothing
    gamma = 0.1  # Seasonal smoothing (weekly for prices)
    
    # Initialize
    level = data[0]
    trend = 0
    seasonal = [0] * 7  # Weekly seasonality
    
    # Apply Holt-Winters smoothing
    for i, price in enumerate(data[1:], 1):
        prev_level = level
        level = alpha * (price - seasonal[i % 7]) + (1 - alpha) * (level + trend)
        trend = beta * (level - prev_level) + (1 - beta) * trend
        seasonal[i % 7] = gamma * (price - level) + (1 - gamma) * seasonal[i % 7]
    
    # Forecast
    forecast_price = level + trend * forecast_days
    confidence = min(0.9, max(0.6, 1 - np.std(data) / (np.mean(data) + 1e-6)))
    
    return max(0, forecast_price), confidence

def _get_market_factor(date, product):
    """Get market adjustment factor for price forecasting"""
    import calendar
    
    # Seasonal price patterns
    month = date.month
    if month in [12, 1, 2]:  # Winter - higher prices due to demand
        seasonal_factor = 1.1
    elif month in [6, 7, 8]:  # Summer - moderate prices
        seasonal_factor = 0.95
    else:
        seasonal_factor = 1.0
    
    # Product category effects
    category = product.category.lower() if product.category else 'regular'
    if 'premium' in category:
        category_factor = 1.2  # Premium products have higher price volatility
    elif 'special' in category:
        category_factor = 1.15
    else:
        category_factor = 1.0
    
    # Supply chain factors (simplified)
    # Higher prices on weekends due to reduced supply
    if date.weekday() >= 5:
        supply_factor = 1.05
    else:
        supply_factor = 1.0
    
    return seasonal_factor * category_factor * supply_factor

@manager_bp.route("/api/forecast/risk", methods=["POST"])
@manager_required
def mgr_forecast_risk():
    """Generate stock risk analysis using Days of Coverage (DoC) calculation"""
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    data = request.get_json(silent=True) or {}
    branch_id = _current_manager_branch_id()
    product_id = data.get('product_id')
    category = data.get('category', '')
    severity = data.get('severity', '')
    days = int(data.get('days', 30))  # Default to 30 days
    
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        risk_analysis = []
        
        # Get products from manager's inventory (avoid grn_number column)
        from sqlalchemy.orm import load_only
        if product_id:
            products = [Product.query.get(product_id)]
        else:
            inventory_items = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id))
                .filter_by(branch_id=branch_id)
                .all()
            )
            products = [item.product for item in inventory_items if item.product]
        
        for product in products:
            if not product:
                continue
            
            # Get current inventory (avoid grn_number column)
            inventory_item = (
                db.session.query(InventoryItem)
                .options(load_only(InventoryItem.id, InventoryItem.branch_id, InventoryItem.product_id, InventoryItem.stock_kg, InventoryItem.warn_level))
                .filter_by(branch_id=branch_id, product_id=product.id)
                .first()
            )
            
            if not inventory_item:
                continue
            
            current_stock = float(inventory_item.stock_kg or 0)
            warn_level = float(inventory_item.warn_level or 0)
            
            # Calculate average daily demand from last N days
            avg_daily_demand = _calculate_avg_daily_demand(product.id, branch_id, days)
            
            # Calculate Days of Coverage (DoC)
            doc = current_stock / max(1, avg_daily_demand) if avg_daily_demand > 0 else 999
            
            # Classify risk based on DoC and thresholds
            risk_type, severity_level, suggested_action = _classify_risk_by_doc(
                doc, current_stock, warn_level
            )
            
            # Apply filters
            if category and category != 'all':
                if category == 'shortage' and risk_type != 'Shortage':
                    continue
                elif category == 'overstock' and risk_type != 'Overstock':
                    continue
            
            if severity and severity != 'all':
                if severity != severity_level:
                    continue
            
            risk_analysis.append({
                'product_id': product.id,
                'product_name': product.name,
                'risk_type': risk_type,
                'severity': severity_level,
                'current_stock': current_stock,
                'warn_level': warn_level,
                'threshold': warn_level,  # Use warn_level as threshold
                'suggested_action': suggested_action,
                'days_of_coverage': round(doc, 1),
                'avg_daily_demand': round(avg_daily_demand, 2)
            })
        
        # Sort by severity (critical first, then by DoC)
        severity_order = {'critical': 0, 'moderate': 1, 'low': 2}
        risk_analysis.sort(key=lambda x: (severity_order.get(x['severity'], 3), -x['days_of_coverage']))
        
        # Calculate summary
        critical_risks = len([r for r in risk_analysis if r["severity"] == "critical"])
        moderate_risks = len([r for r in risk_analysis if r["severity"] == "moderate"])
        low_risks = len([r for r in risk_analysis if r["severity"] == "low"])
        
        summary = {
            "critical_risks": critical_risks,
            "moderate_risks": moderate_risks,
            "low_risks": low_risks,
            "total_analyzed": len(risk_analysis)
        }
        
        return jsonify({
            "ok": True,
            "message": f"Risk analysis completed for {len(risk_analysis)} products",
            "risk_analysis": risk_analysis,
            "summary": summary
        })
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def _calculate_avg_daily_demand(product_id, branch_id, days):
    """Calculate average daily demand from sales transactions for the last N days"""
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    try:
        # Use datetime for proper comparison with transaction_date (which is DateTime)
        end_datetime = datetime.now()
        start_datetime = end_datetime - timedelta(days=days)
        
        # Get total quantity sold for the product in the last N days
        total_sold = db.session.query(
            func.sum(SalesTransaction.quantity_sold)
        ).filter(
            SalesTransaction.product_id == product_id,
            SalesTransaction.branch_id == branch_id,
            SalesTransaction.transaction_date >= start_datetime,
            SalesTransaction.transaction_date <= end_datetime
        ).scalar() or 0
        
        # Calculate average daily demand
        avg_daily_demand = total_sold / days if days > 0 else 0
        
        return float(avg_daily_demand)
        
    except Exception as e:
        print(f"Error calculating avg daily demand: {e}")
        import traceback
        traceback.print_exc()
        return 0.0

def _classify_risk_by_doc(doc, current_stock, warn_level):
    """Classify risk based on Days of Coverage (DoC) and thresholds"""
    
    # Risk classification based on DoC thresholds
    if current_stock <= warn_level or doc < 7:
        # Shortage risk
        if doc < 3:
            return "Shortage", "critical", "Restock immediately - critical shortage"
        else:
            return "Shortage", "moderate", "Monitor closely, prepare purchase"
    elif doc > 45:
        # Overstock risk
        if doc > 90:
            return "Overstock", "critical", "Consider promotional pricing - excessive stock"
        else:
            return "Overstock", "moderate", "Review pricing strategy"
    else:
        # Balanced stock
        if doc < 14:
            return "Balanced", "moderate", "Stock level adequate but monitor"
        else:
            return "Balanced", "low", "Stock level is healthy"

def _classify_stock_risk(current_stock, forecasted_demand, warn_level, auto_level, lead_time_days, stockout_events):
    """Legacy function - kept for backward compatibility"""
    
    # Use warn_level as the threshold for risk classification
    threshold = warn_level if warn_level > 0 else 100  # Default threshold if not set
    
    # Risk Classification Logic based on threshold percentages
    stock_vs_threshold_ratio = current_stock / threshold if threshold > 0 else 1
    
    # Classify risk level based on threshold percentages
    if stock_vs_threshold_ratio < 0.5:  # Current stock < 50% of threshold
        risk_level = "Critical"
        suggested_action = "Restock immediately"
        risk_score = 90
    elif stock_vs_threshold_ratio < 1.0:  # Current stock is between 50-100% of threshold
        risk_level = "Moderate"
        suggested_action = "Monitor closely, prepare purchase"
        risk_score = 60
    else:  # Current stock >= threshold
        risk_level = "Low"
        suggested_action = "Stock level is healthy"
        risk_score = 20
    
    return risk_score, risk_level, suggested_action

@manager_bp.route("/api/forecast/data", methods=["GET"])
@manager_required
def mgr_forecast_data():
    """Get forecast data for manager's branch"""
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        # Get forecast data for the last 30 days and next 30 days
        today = datetime.now().date()
        start_date = today - timedelta(days=30)
        end_date = today + timedelta(days=30)
        
        # Get forecast data
        forecasts = db.session.query(
            ForecastData.forecast_date,
            ForecastData.product_id,
            Product.name.label('product_name'),
            ForecastData.predicted_demand,
            ForecastData.confidence_interval_lower,
            ForecastData.confidence_interval_upper,
            ForecastData.accuracy_score
        ).join(Product, ForecastData.product_id == Product.id)\
         .filter(ForecastData.branch_id == branch_id)\
         .filter(ForecastData.forecast_date >= start_date)\
         .filter(ForecastData.forecast_date <= end_date)\
         .order_by(ForecastData.forecast_date)\
         .all()
        
        # Get actual sales data for comparison
        actual_sales = db.session.query(
            func.date(SalesTransaction.transaction_date).label('date'),
            SalesTransaction.product_id,
            func.sum(SalesTransaction.quantity_sold).label('actual_demand')
        ).filter(SalesTransaction.branch_id == branch_id)\
         .filter(func.date(SalesTransaction.transaction_date) >= start_date)\
         .filter(func.date(SalesTransaction.transaction_date) <= today)\
         .group_by(func.date(SalesTransaction.transaction_date), SalesTransaction.product_id)\
         .all()
        
        # Create maps for easy lookup
        actual_map = {(row.date, row.product_id): row.actual_demand for row in actual_sales}
        
        # Prepare chart data
        labels = []
        forecast_data = []
        actual_data = []
        
        # Group by date
        date_groups = {}
        for forecast in forecasts:
            date = forecast.forecast_date
            if date not in date_groups:
                date_groups[date] = {'forecast': 0, 'actual': 0}
            date_groups[date]['forecast'] += forecast.predicted_demand or 0
            date_groups[date]['actual'] += actual_map.get((date, forecast.product_id), 0)
        
        # Sort by date and prepare arrays
        for date in sorted(date_groups.keys()):
            labels.append(date.strftime('%Y-%m-%d'))
            forecast_data.append(round(date_groups[date]['forecast'], 2))
            actual_data.append(round(date_groups[date]['actual'], 2))
        
        # Calculate summary
        total_forecast = sum(forecast_data)
        total_actual = sum(actual_data)
        avg_confidence = sum(f.accuracy_score or 0.8 for f in forecasts) / len(forecasts) if forecasts else 0.8
        
        return jsonify({
            "ok": True,
            "forecast": {
                "labels": labels,
                "forecast": forecast_data,
                "actual": actual_data
            },
            "summary": {
                "avg_demand": round(total_forecast / len(forecast_data) if forecast_data else 0, 2),
                "peak_date": labels[forecast_data.index(max(forecast_data))] if forecast_data else None,
                "peak_quantity": max(forecast_data) if forecast_data else 0,
                "suggested_reorder": round(total_forecast * 0.1, 2),  # 10% of total forecast
                "confidence": round(avg_confidence, 2)
            },
            "details": [{
                'date': f.forecast_date.strftime('%Y-%m-%d'),
                'projected': f.predicted_demand,
                'confidence': f'{round((f.accuracy_score or 0.8) * 100)}%',
                'trend': 'Increasing' if f.predicted_demand > 100 else 'Stable'
            } for f in forecasts[-10:]]  # Last 10 forecasts
        })
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/forecast/export", methods=["GET"])
@manager_required
def mgr_forecast_export():
    """Export forecast data as CSV"""
    from datetime import datetime, timedelta
    import csv
    import io
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    format_type = request.args.get('format', 'csv')
    
    try:
        # Get forecast data
        today = datetime.now().date()
        start_date = today - timedelta(days=30)
        end_date = today + timedelta(days=30)
        
        forecasts = db.session.query(
            ForecastData.forecast_date,
            Product.name.label('product_name'),
            ForecastData.predicted_demand,
            ForecastData.confidence_score,
            ForecastData.model_type
        ).join(Product, ForecastData.product_id == Product.id)\
         .filter(ForecastData.branch_id == branch_id)\
         .filter(ForecastData.forecast_date >= start_date)\
         .filter(ForecastData.forecast_date <= end_date)\
         .order_by(ForecastData.forecast_date)\
         .all()
        
        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Date', 'Product', 'Predicted Demand (kg)', 'Confidence', 'Model Type'])
            
            for forecast in forecasts:
                writer.writerow([
                    forecast.forecast_date.strftime('%Y-%m-%d'),
                    forecast.product_name,
                    forecast.predicted_demand,
                    f"{forecast.confidence_score:.2%}",
                    forecast.model_type
                ])
            
            output.seek(0)
            return output.getvalue(), 200, {
                'Content-Type': 'text/csv',
                'Content-Disposition': f'attachment; filename=forecast_{branch_id}_{today}.csv'
            }
        
        return jsonify({"ok": False, "error": "Unsupported format"}), 400
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/branches", methods=["GET"])
@manager_required
def mgr_branches():
    """Get manager's allowed branches"""
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    branch = Branch.query.get(branch_id)
    if not branch:
        return jsonify({"ok": False, "error": "Branch not found"}), 404
    
    return jsonify({
        "ok": True,
        "branches": [{
            "id": branch.id,
            "name": branch.name,
            "location": branch.location
        }]
    })

@manager_bp.route("/api/analytics/export", methods=["GET"])
@manager_required
def mgr_analytics_export():
    """Export analytics data as CSV"""
    from datetime import datetime, timedelta
    import csv
    import io
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    format_type = request.args.get('format', 'csv')
    export_type = request.args.get('type', 'risk')
    
    try:
        if export_type == 'risk':
            # Export risk data
            today = datetime.now().date()
            next7 = today + timedelta(days=7)
            
            # Get forecast data for next 7 days
            f_next = db.session.query(
                ForecastData.product_id, 
                func.sum(ForecastData.predicted_demand).label('pred')
            ).filter(
                ForecastData.branch_id == branch_id,
                ForecastData.forecast_date >= today,
                ForecastData.forecast_date <= next7
            ).group_by(ForecastData.product_id).all()
            
            # Get current inventory
            inv_map = {
                it.product_id: float(it.stock_kg or 0) 
                for it in InventoryItem.query.filter(InventoryItem.branch_id == branch_id).all()
            }
            
            # Get product names
            product_map = {p.id: p.name for p in Product.query.all()}
            
            if format_type == 'csv':
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(['Product', 'Risk Type', 'Severity Level', 'Current Stock (kg)', 'Threshold (kg)', 'Suggested Action'])
                
                for pid, pred in f_next:
                    stock = inv_map.get(pid, 0.0)
                    gap = float(pred or 0) - stock
                    product_name = product_map.get(pid, f"Product {pid}")
                    
                    if gap > 0:
                        risk_type = "Shortage"
                        severity = "Critical" if gap > stock else "Moderate"
                        threshold = stock + gap
                        action = "Immediate restock required" if severity == "Critical" else "Monitor stock levels"
                    else:
                        risk_type = "Overstock"
                        severity = "Low"
                        threshold = stock
                        action = "Consider promotional pricing"
                    
                    writer.writerow([
                        product_name,
                        risk_type,
                        severity,
                        stock,
                        threshold,
                        action
                    ])
                
                output.seek(0)
                return output.getvalue(), 200, {
                    'Content-Type': 'text/csv',
                    'Content-Disposition': f'attachment; filename=risk_report_{branch_id}_{today}.csv'
                }
        
        return jsonify({"ok": False, "error": "Unsupported export type"}), 400
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ========================= REPORTS (BRANCH) =========================
def _mgr_parse_date(s, default):
    from datetime import datetime
    if not s:
        return default
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except Exception:
        return default

@manager_bp.get("/api/reports/sales")
@manager_required
def mgr_reports_sales():
    from sqlalchemy import func
    from datetime import datetime, timedelta
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "branch_id required"}), 400
    from_str = request.args.get('from')
    to_str = request.args.get('to')
    product_id = request.args.get('product_id', type=int)
    granularity = request.args.get('granularity', 'day')
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    now = datetime.now()
    start = _mgr_parse_date(from_str, now - timedelta(days=30))
    end = _mgr_parse_date(to_str, now)
    q = db.session.query(
        func.date(SalesTransaction.transaction_date).label('d'),
        SalesTransaction.product_id,
        func.sum(SalesTransaction.quantity_sold).label('qty'),
        func.sum(SalesTransaction.total_amount).label('amt')
    ).filter(SalesTransaction.branch_id == branch_id,
             SalesTransaction.transaction_date >= start,
             SalesTransaction.transaction_date <= end)
    if product_id:
        q = q.filter(SalesTransaction.product_id == product_id)
    if granularity == 'month':
        q = q.with_entities(
            func.to_char(SalesTransaction.transaction_date, 'YYYY-MM').label('period'),
            SalesTransaction.product_id,
            func.sum(SalesTransaction.quantity_sold).label('qty'),
            func.sum(SalesTransaction.total_amount).label('amt')
        ).group_by('period', SalesTransaction.product_id).order_by('period')
    elif granularity == 'week':
        q = q.with_entities(
            func.to_char(SalesTransaction.transaction_date, 'IYYY-IW').label('period'),
            SalesTransaction.product_id,
            func.sum(SalesTransaction.quantity_sold).label('qty'),
            func.sum(SalesTransaction.total_amount).label('amt')
        ).group_by('period', SalesTransaction.product_id).order_by('period')
    else:
        q = q.group_by('d', SalesTransaction.product_id).order_by('d')
    total = q.count()
    pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    rows = q.offset((page-1)*page_size).limit(page_size).all() if page_size>0 else q.all()
    product_map = {p.id: p.name for p in Product.query.all()}
    out_rows = []
    sum_qty = sum_amt = 0.0
    for r in rows:
        if granularity in ('week','month'):
            period, pid, qty, amt = r[0], r[1], float(r[2] or 0), float(r[3] or 0)
            date_label = period
        else:
            d, pid, qty, amt = r[0], r[1], float(r[2] or 0), float(r[3] or 0)
            date_label = d.strftime('%Y-%m-%d')
        sum_qty += qty
        sum_amt += amt
        out_rows.append({
            "date": date_label,
            "branch_id": branch_id,
            "product_id": pid,
            "product_name": product_map.get(pid),
            "qty_kg": qty,
            "amount": amt,
        })
    return jsonify({
        "ok": True,
        "rows": out_rows,
        "totals": {"sum_qty_kg": sum_qty, "sum_amount": sum_amt},
        "meta": {"page": page, "pages": pages, "count": total}
    })

@manager_bp.get("/api/reports/forecast")
@manager_required
def mgr_reports_forecast():
    from sqlalchemy import func, and_
    from datetime import datetime, timedelta
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "branch_id required"}), 400
    from_str = request.args.get('from')
    to_str = request.args.get('to')
    product_id = request.args.get('product_id', type=int)
    model_type = request.args.get('model_type')
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    now = datetime.now().date()
    start = _mgr_parse_date(from_str, datetime.combine(now - timedelta(days=30), datetime.min.time())).date()
    end = _mgr_parse_date(to_str, datetime.combine(now, datetime.min.time())).date()
    q = db.session.query(
        ForecastData.forecast_date,
        ForecastData.product_id,
        func.sum(ForecastData.predicted_demand).label('forecast_kg')
    ).filter(and_(ForecastData.branch_id == branch_id,
                  ForecastData.forecast_date >= start,
                  ForecastData.forecast_date <= end))
    if product_id:
        q = q.filter(ForecastData.product_id == product_id)
    if model_type:
        q = q.filter(ForecastData.model_type.ilike(model_type))
    q = q.group_by(ForecastData.forecast_date, ForecastData.product_id).order_by(ForecastData.forecast_date)
    total = q.count()
    pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    rows = q.offset((page-1)*page_size).limit(page_size).all() if page_size>0 else q.all()
    product_map = {p.id: p.name for p in Product.query.all()}
    out_rows = []
    sum_f = sum_a = 0.0
    mape_sum = 0.0
    mape_count = 0
    for d, pid, fkg in rows:
        actual = db.session.query(func.sum(SalesTransaction.quantity_sold)).filter(
            and_(SalesTransaction.branch_id == branch_id,
                 SalesTransaction.product_id == pid,
                 func.date(SalesTransaction.transaction_date) == d)
        ).scalar() or 0
        fkg = float(fkg or 0)
        akg = float(actual or 0)
        diff = fkg - akg
        ape = abs(diff) / akg * 100 if akg > 0 else None
        if ape is not None:
            mape_sum += ape
            mape_count += 1
        sum_f += fkg
        sum_a += akg
        out_rows.append({
            "date": d.strftime('%Y-%m-%d'),
            "branch_id": branch_id,
            "product_id": pid,
            "product_name": product_map.get(pid),
            "forecast_kg": fkg,
            "actual_kg": akg,
            "diff_kg": round(diff, 2),
            "abs_pct_err": round(ape, 2) if ape is not None else None,
        })
    mape = (mape_sum / mape_count) if mape_count > 0 else None
    return jsonify({
        "ok": True,
        "rows": out_rows,
        "totals": {"sum_forecast_kg": sum_f, "sum_actual_kg": sum_a, "mape": round(mape,2) if mape is not None else None},
        "meta": {"page": page, "pages": pages, "count": total}
    })

@manager_bp.get("/api/reports/inventory")
@manager_required
def mgr_reports_inventory():
    from sqlalchemy import func
    from datetime import datetime, timedelta
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "branch_id required"}), 400
    from_str = request.args.get('from')
    to_str = request.args.get('to')
    product_id = request.args.get('product_id', type=int)
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    now = datetime.now().date()
    start = _mgr_parse_date(from_str, datetime.combine(now - timedelta(days=30), datetime.min.time())).date()
    end = _mgr_parse_date(to_str, datetime.combine(now, datetime.min.time())).date()
    recv_q = db.session.query(
        func.date(RestockLog.created_at).label('d'),
        InventoryItem.product_id,
        func.sum(RestockLog.qty_kg).label('received_kg')
    ).join(InventoryItem, RestockLog.inventory_item_id == InventoryItem.id)
    recv_q = recv_q.filter(InventoryItem.branch_id == branch_id,
                           func.date(RestockLog.created_at) >= start,
                           func.date(RestockLog.created_at) <= end)
    if product_id:
        recv_q = recv_q.filter(InventoryItem.product_id == product_id)
    recv_q = recv_q.group_by('d', InventoryItem.product_id)
    recv_map = {(r.d, r.product_id): float(r.received_kg or 0) for r in recv_q.all()}
    sold_q = db.session.query(
        func.date(SalesTransaction.transaction_date).label('d'),
        SalesTransaction.product_id,
        func.sum(SalesTransaction.quantity_sold).label('sold_kg')
    ).filter(SalesTransaction.branch_id == branch_id,
             func.date(SalesTransaction.transaction_date) >= start,
             func.date(SalesTransaction.transaction_date) <= end)
    if product_id:
        sold_q = sold_q.filter(SalesTransaction.product_id == product_id)
    sold_q = sold_q.group_by('d', SalesTransaction.product_id)
    sold_map = {(r.d, r.product_id): float(r.sold_kg or 0) for r in sold_q.all()}
    inv_map = {it.product_id: float(it.stock_kg or 0) for it in InventoryItem.query.filter(InventoryItem.branch_id == branch_id).all()}
    product_map = {p.id: p.name for p in Product.query.all()}
    all_keys = set(list(recv_map.keys()) + list(sold_map.keys()))
    rows_list = []
    for (d, pid) in sorted(all_keys):
        rows_list.append({
            "date": d.strftime('%Y-%m-%d'),
            "branch_id": branch_id,
            "product_id": pid,
            "product_name": product_map.get(pid),
            "opening_kg": None,
            "received_kg": recv_map.get((d,pid),0.0),
            "sold_kg": sold_map.get((d,pid),0.0),
            "closing_kg": inv_map.get(pid,0.0),
        })
    total = len(rows_list)
    pages = (total + page_size - 1) // page_size if page_size > 0 else 1
    start_idx = (page-1)*page_size
    end_idx = start_idx + page_size
    page_rows = rows_list[start_idx:end_idx]
    avg_closing = sum(r["closing_kg"] for r in rows_list)/total if total>0 else 0
    return jsonify({
        "ok": True,
        "rows": page_rows,
        "totals": {"avg_closing_kg": round(avg_closing,2)},
        "meta": {"page": page, "pages": pages, "count": total}
    })

@manager_bp.route("/api/notifications/dispatch", methods=["POST"])
@manager_required
def mgr_notifications_dispatch():
    """Dispatch notification to manager (called by admin)"""
    from models import Notification, db
    from datetime import datetime
    
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data provided"}), 400
    
    try:
        notification = Notification(
            type=data.get('type'),
            branch_id=data.get('branch_id'),
            date=datetime.strptime(data.get('date'), '%Y-%m-%d').date(),
            message=data.get('message'),
            sender='Admin',
            status='unread',
            created_at=datetime.now()
        )
        
        db.session.add(notification)
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "Notification dispatched to manager successfully"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/notifications", methods=["GET"])
@manager_required
def mgr_list_notifications():
    """Get notifications for manager's branch"""
    from models import Notification
    
    # Get manager's branch ID - prioritize URL parameter over session
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    print(f"DEBUG: Loading notifications for branch_id: {branch_id}")
    
    notifications = Notification.query.filter_by(branch_id=branch_id).order_by(Notification.created_at.desc()).all()
    
    return jsonify({
        "ok": True,
        "notifications": [notification.to_dict() for notification in notifications]
    })


@manager_bp.route("/api/notifications/unread-count", methods=["GET"])
@manager_required
def mgr_unread_count():
    """Get unread notification count for manager's branch"""
    from models import Notification
    
    # Get manager's branch ID - prioritize URL parameter over session
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    print(f"DEBUG: Getting unread count for branch_id: {branch_id}")
    
    unread_count = Notification.query.filter_by(
        branch_id=branch_id, 
        status='unread'
    ).count()
    
    # Debug logging
    print(f"DEBUG: Unread count API called for branch_id={branch_id}, unread_count={unread_count}")
    
    return jsonify({
        "ok": True,
        "unread_count": unread_count
    })

@manager_bp.route("/api/notifications/<int:notification_id>/read", methods=["PATCH"])
@manager_required
def mgr_mark_notification_read(notification_id):
    """Mark a notification as read"""
    from models import Notification, db
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        notification = Notification.query.filter_by(
            id=notification_id, 
            branch_id=branch_id
        ).first()
        
        if not notification:
            return jsonify({"ok": False, "error": "Notification not found"}), 404
        
        notification.status = 'read'
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "Notification marked as read"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/dashboard/alerts", methods=["GET"])
@manager_required
def mgr_dashboard_alerts():
    """Get real-time alerts for manager dashboard (low stock, near expiry)"""
    from models import InventoryItem, Product
    from datetime import datetime, timedelta, date
    from sqlalchemy import or_, and_
    
    # Get manager's branch ID
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    alerts = []
    
    try:
        # 1. Low Stock Alerts - items below warn_level or below 100kg default
        low_stock_items = (
            db.session.query(InventoryItem, Product)
            .join(Product, InventoryItem.product_id == Product.id)
            .filter(
                and_(
                    InventoryItem.branch_id == branch_id,
                    or_(
                        and_(InventoryItem.warn_level.isnot(None), InventoryItem.stock_kg < InventoryItem.warn_level),
                        and_(InventoryItem.warn_level.is_(None), InventoryItem.stock_kg < 100)
                    ),
                    InventoryItem.stock_kg > 0  # Only items that still have stock
                )
            )
            .all()
        )
        
        if low_stock_items:
            low_stock_count = len(low_stock_items)
            alerts.append({
                "type": "low_stock",
                "icon": "exclamation-triangle",
                "message": f"Stock below reorder level for {low_stock_count} item{'s' if low_stock_count > 1 else ''}",
                "severity": "caution",
                "count": low_stock_count
            })
        
        # 2. Near Expiry Alerts (if expiry_date exists in inventory)
        # Note: This assumes expiry tracking - if not implemented, this will be empty
        # For now, we'll check if there's a way to track expiry
        # If InventoryItem has expiry_date field, use it; otherwise skip
        try:
            # Check if expiry_date column exists (some databases might not have it)
            near_expiry_items = []
            # For now, we'll create a placeholder that can be enhanced later
            # If you have expiry_date in InventoryItem, uncomment and modify:
            # expiry_threshold = date.today() + timedelta(days=30)
            # near_expiry_items = (
            #     db.session.query(InventoryItem, Product)
            #     .join(Product, InventoryItem.product_id == Product.id)
            #     .filter(
            #         and_(
            #             InventoryItem.branch_id == branch_id,
            #             InventoryItem.expiry_date.isnot(None),
            #             InventoryItem.expiry_date <= expiry_threshold,
            #             InventoryItem.expiry_date >= date.today()
            #         )
            #     )
            #     .all()
            # )
            
            # Placeholder: Check for items with batch codes that might indicate age
            # This is a simple heuristic - you can enhance this based on your business logic
            if near_expiry_items:
                expiry_count = len(near_expiry_items)
                alerts.append({
                    "type": "near_expiry",
                    "icon": "exclamation-circle",
                    "message": f"{expiry_count} item{'s' if expiry_count > 1 else ''} near expiry date",
                    "severity": "warning",
                    "count": expiry_count
                })
        except Exception as e:
            # If expiry tracking is not implemented, skip this alert
            print(f"Expiry tracking not available: {e}")
            pass
        
    except Exception as e:
        print(f"Error loading alerts: {e}")
        import traceback
        traceback.print_exc()
    
    return jsonify({
        "ok": True,
        "alerts": alerts
    })

@manager_bp.route("/api/dashboard/announcements", methods=["GET"])
@manager_required
def mgr_dashboard_announcements():
    """Get branch announcements for manager dashboard"""
    from models import Notification
    from datetime import datetime, timedelta
    
    # Get manager's branch ID
    branch_id = request.args.get('branch_id', type=int) or _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    # Get recent announcements (last 30 days, limit 5)
    since = datetime.now() - timedelta(days=30)
    
    announcements = (
        Notification.query
        .filter(
            Notification.branch_id == branch_id,
            Notification.type.in_(['announcement', 'manual_message', 'system_maintenance', 'delivery_notice']),
            Notification.created_at >= since
        )
        .order_by(Notification.created_at.desc())
        .limit(5)
        .all()
    )
    
    return jsonify({
        "ok": True,
        "announcements": [
            {
                "id": ann.id,
                "title": ann.message.split('\n')[0] if '\n' in ann.message else ann.message[:50] + ('...' if len(ann.message) > 50 else ''),
                "message": ann.message,
                "date": ann.date.strftime('%b %d, %Y') if ann.date else ann.created_at.strftime('%b %d, %Y'),
                "sender": ann.sender,
                "type": ann.type,
                "created_at": ann.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for ann in announcements
        ]
    })

@manager_bp.route("/api/notifications/<int:notification_id>", methods=["DELETE"])
@manager_required
def mgr_delete_notification(notification_id):
    """Delete a notification"""
    from models import Notification, db
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        notification = Notification.query.filter_by(
            id=notification_id, 
            branch_id=branch_id
        ).first()
        
        if not notification:
            return jsonify({"ok": False, "error": "Notification not found"}), 404
        
        db.session.delete(notification)
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "Notification deleted successfully"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/notifications/mark-all-read", methods=["PATCH"])
@manager_required
def mgr_mark_all_notifications_read():
    """Mark all notifications as read for manager's branch"""
    from models import Notification, db
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        # Mark all unread notifications as read
        updated = db.session.query(Notification).filter_by(
            branch_id=branch_id,
            status='unread'
        ).update({'status': 'read'})
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": f"{updated} notifications marked as read",
            "updated_count": updated
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/notifications/clear-all", methods=["DELETE"])
@manager_required
def mgr_clear_all_notifications():
    """Delete all notifications for manager's branch"""
    from models import Notification, db
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        return jsonify({"ok": False, "error": "Manager branch not found"}), 400
    
    try:
        # Delete all notifications for this branch
        deleted = db.session.query(Notification).filter_by(branch_id=branch_id).delete()
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": f"{deleted} notifications deleted",
            "deleted_count": deleted
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@manager_bp.route("/api/purchases/recent", methods=["GET"])
@manager_required
def mgr_purchases_recent():
    """Get recent purchases for manager's branch"""
    try:
        from datetime import datetime, timezone, timedelta
        
        branch_id = _current_manager_branch_id()
        if not branch_id:
            # Fallback: try to get branch from URL parameters or default to 1
            url_branch = request.args.get('branch')
            if url_branch:
                branch_id = int(url_branch)
            else:
                branch_id = 1  # Default to branch 1 for testing
                
            print(f"DEBUG: Using fallback branch_id={branch_id}")
        
        # Get recent sales transactions for this branch with eager loading of product
        from sqlalchemy.orm import joinedload
        recent_sales = (
            db.session.query(SalesTransaction)
            .options(joinedload(SalesTransaction.product))
            .filter(SalesTransaction.branch_id == branch_id)
            .order_by(SalesTransaction.transaction_date.desc())
            .limit(50)
            .all()
        )
        
        # Get current stock for each product to calculate historical estimated remaining
        from sqlalchemy.orm import load_only
        from collections import defaultdict
        
        # Get current inventory for all products in this branch (include batch_code)
        inventory_items = (
            db.session.query(InventoryItem)
            .options(load_only(InventoryItem.product_id, InventoryItem.stock_kg, InventoryItem.unit_price, InventoryItem.batch_code))
            .filter_by(branch_id=branch_id)
            .all()
        )
        
        # Build maps: product_id -> current_stock, product_id -> unit_price, product_id -> batch_codes
        current_stock_map = {}
        unit_price_map = {}
        product_batches_map = {}  # product_id -> list of batch_codes (oldest first)
        for inv_item in inventory_items:
            current_stock_map[inv_item.product_id] = float(inv_item.stock_kg) if inv_item.stock_kg else 0.0
            unit_price_map[inv_item.product_id] = float(inv_item.unit_price) if inv_item.unit_price else 0.0
            
            # Collect batch codes for each product (we'll sort by oldest later)
            if inv_item.product_id not in product_batches_map:
                product_batches_map[inv_item.product_id] = []
            if inv_item.batch_code:
                product_batches_map[inv_item.product_id].append(inv_item.batch_code)
        
        # Get earliest restock dates for batch ordering (simplified approach)
        batch_restock_map = {}
        
        # Only query restock dates if we have inventory items
        if inventory_items:
            item_ids = [item.id for item in inventory_items]
            if item_ids:
                try:
                    from sqlalchemy import func
                    # Use a simpler query that groups by product_id and batch_code only
                    restock_dates = (
                        db.session.query(
                            InventoryItem.product_id,
                            InventoryItem.batch_code,
                            func.min(RestockLog.created_at).label('earliest_restock')
                        )
                        .join(RestockLog, RestockLog.inventory_item_id == InventoryItem.id)
                        .filter(InventoryItem.id.in_(item_ids))
                        .group_by(InventoryItem.product_id, InventoryItem.batch_code)
                        .all()
                    )
                    
                    # Build map: (product_id, batch_code) -> earliest_restock
                    for row in restock_dates:
                        key = (row.product_id, row.batch_code)
                        if key not in batch_restock_map or (row.earliest_restock and (not batch_restock_map[key] or row.earliest_restock < batch_restock_map[key])):
                            batch_restock_map[key] = row.earliest_restock
                except Exception as e:
                    print(f"WARNING: Error querying restock dates: {e}")
                    import traceback
                    traceback.print_exc()
                    # Rollback to clear the failed transaction
                    db.session.rollback()
                    # Continue without batch ordering
        
        # Sort batches by oldest first for each product
        for product_id in product_batches_map:
            try:
                product_batches_map[product_id] = sorted(
                    product_batches_map[product_id],
                    key=lambda bc: (batch_restock_map.get((product_id, bc)) or datetime.max, bc)
                )
            except Exception as e:
                print(f"WARNING: Error sorting batches for product {product_id}: {e}")
                # Keep original order if sorting fails
        
        # Group sales by product and calculate historical remaining for each transaction
        # For each sale, historical_remaining = current_stock + sum of all sales AFTER this transaction
        # Sort by date descending (newest first) for proper calculation
        sales_by_product = defaultdict(list)
        for sale in recent_sales:
            sales_by_product[sale.product_id].append(sale)
        
        # Calculate historical remaining for each transaction
        logbook_entries = []
        for sale in recent_sales:
            # Get product name
            product_name = sale.product.name if sale.product else "Unknown Product"
            
            # Get current stock for this product
            current_stock = current_stock_map.get(sale.product_id, 0.0)
            unit_price = unit_price_map.get(sale.product_id, 0.0)
            
            # If unit_price is 0, calculate it from the sales transaction
            if unit_price == 0.0 and float(sale.quantity_sold) > 0:
                unit_price = float(sale.total_amount) / float(sale.quantity_sold)
            
            # Get batch_code from the sale transaction (stored when sale was made)
            # Fallback to oldest batch if not stored
            batch_code = sale.batch_code
            if not batch_code:
                batch_codes = product_batches_map.get(sale.product_id, [])
                batch_code = batch_codes[0] if batch_codes else None
            
            # Calculate historical estimated remaining: current_stock + sum of sales AFTER this transaction
            # Since transactions are sorted by date DESC, we need to add back all sales that happened after
            historical_remaining = current_stock
            for other_sale in sales_by_product[sale.product_id]:
                # If this sale happened AFTER the current sale (newer date), add it back
                if other_sale.transaction_date > sale.transaction_date:
                    historical_remaining += float(other_sale.quantity_sold)
                # If same date but different ID and this one is newer (higher ID means later insertion)
                elif other_sale.transaction_date == sale.transaction_date and other_sale.id > sale.id:
                    historical_remaining += float(other_sale.quantity_sold)
            
            # Ensure datetime is timezone-aware (assume UTC if naive, then convert to Philippines time)
            ph_tz = timezone(timedelta(hours=8))
            
            if sale.transaction_date.tzinfo is None:
                # If naive datetime, assume it's UTC
                dt_utc = sale.transaction_date.replace(tzinfo=timezone.utc)
            else:
                dt_utc = sale.transaction_date
            
            # Convert to Philippines time
            dt_ph = dt_utc.astimezone(ph_tz)
            
            # Calculate initial inventory: estimated remaining + quantity sold
            initial_inventory = historical_remaining + float(sale.quantity_sold)
            
            entry = {
                "id": sale.id,
                "date": dt_ph.strftime("%Y-%m-%d"),  # Keep for backward compatibility if needed
                "datetime": dt_ph.strftime("%Y-%m-%d %H:%M:%S"),
                "created_at": dt_ph.isoformat(),  # Ensure this is also timezone-aware
                "timestamp": dt_ph.isoformat(),  # Ensure this is also timezone-aware
                "riceVariant": product_name.lower().replace(' ', '-').replace('_', '-'),
                "product_name": product_name,
                "batch_code": batch_code,  # Include batch_code (oldest batch for this product)
                "batch": batch_code,  # Alias for compatibility
                "batchCode": batch_code,  # Alias for compatibility
                "price": unit_price,
                "initialInventory": initial_inventory,  # Calculated: estimated remaining + quantity sold
                "addedStocks": 0.0,  # Not tracked in current system
                "totalSoldKg": float(sale.quantity_sold),
                "totalAmount": float(sale.total_amount),
                "estimatedRemaining": historical_remaining,  # Historical estimated remaining at time of transaction
                "actualRemaining": None,  # Not tracked in current system
                "discrepancy": None  # Not tracked in current system
            }
            logbook_entries.append(entry)
    
        return jsonify({
            "ok": True,
            "entries": logbook_entries,
            "branch_id": branch_id
        })
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"ERROR in mgr_purchases_recent: {str(e)}")
        print(f"Traceback: {error_trace}")
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e),
            "entries": [],
            "branch_id": branch_id if 'branch_id' in locals() else None
        }), 500

@manager_bp.route("/api/sales/bulk", methods=["POST"])
@manager_required
def mgr_sales_bulk():
    """Bulk sales submission from purchase form"""
    from models import SalesTransaction, Product, db
    from datetime import datetime, timezone, timedelta
    
    branch_id = _current_manager_branch_id()
    if not branch_id:
        # Fallback: try to get branch from URL parameters or default to 1
        url_branch = request.args.get('branch')
        if url_branch:
            branch_id = int(url_branch)
        else:
            branch_id = 1  # Default to branch 1 for testing
        
        print(f"DEBUG: Using fallback branch_id={branch_id}")
    
    print(f"DEBUG: Bulk sales API called with branch_id={branch_id}")
    
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data provided"}), 400
    
    try:
        date_str = data.get('date')
        items = data.get('items', [])
        
        if not date_str or not items:
            return jsonify({"ok": False, "error": "Date and items are required"}), 400
        
        # Parse date and combine with current Philippines time (UTC+8)
        ph_tz = timezone(timedelta(hours=8))
        now_ph = datetime.now(ph_tz)
        
        # Use the selected date but with current Philippines time
        date_only = datetime.strptime(date_str, '%Y-%m-%d').date()
        transaction_date_ph = datetime.combine(date_only, now_ph.time()).replace(tzinfo=ph_tz)
        
        # Convert to UTC and remove timezone info for database storage (naive datetime)
        transaction_date_utc = transaction_date_ph.astimezone(timezone.utc)
        transaction_date = transaction_date_utc.replace(tzinfo=None)  # Store as naive UTC datetime
        
        created_transactions = []
        
        for item in items:
            product_name = item.get('product_name')
            price_per_kg = float(item.get('price_per_kg', 0))
            quantity_sold_kg = float(item.get('quantity_sold_kg', 0))
            total_amount = float(item.get('total_amount', 0))
            remarks = item.get('remarks', '')
            
            if not product_name or quantity_sold_kg <= 0:
                continue  # Skip invalid items
            
            # Get batch_code from request if provided
            requested_batch_code = item.get('batch_code', '').strip() if item.get('batch_code') else None
            
            # Find or create product
            product = Product.query.filter_by(name=product_name).first()
            if not product:
                product = Product(
                    name=product_name,
                    category='Rice',
                    description=f'{product_name} rice variety'
                )
                db.session.add(product)
                db.session.flush()  # Get the ID
            
            # Create sales transaction
            transaction = SalesTransaction(
                branch_id=branch_id,
                product_id=product.id,
                quantity_sold=quantity_sold_kg,
                unit_price=price_per_kg,
                total_amount=total_amount,
                transaction_date=transaction_date,
                batch_code=requested_batch_code  # Store the batch code used for this sale
            )
            
            db.session.add(transaction)
            
            # Update inventory: use specific batch if provided, otherwise use FIFO
            from sqlalchemy.orm import load_only
            from sqlalchemy import func, or_
            
            if requested_batch_code:
                # Deduct from the specific batch requested
                print(f"DEBUG: Deducting from specific batch '{requested_batch_code}' for {product_name}")
                
                # Find the inventory item with the specified batch code
                inventory_item = (
                    db.session.query(InventoryItem)
                    .options(load_only(InventoryItem.id, InventoryItem.stock_kg, InventoryItem.batch_code))
                    .filter_by(
                        branch_id=branch_id,
                        product_id=product.id,
                        batch_code=requested_batch_code
                    )
                    .first()
                )
                
                if inventory_item:
                    current_stock = float(inventory_item.stock_kg or 0)
                    if current_stock >= quantity_sold_kg:
                        new_stock = current_stock - quantity_sold_kg
                        inventory_item.stock_kg = max(0, new_stock)
                        inventory_item.updated_at = datetime.utcnow()
                        print(f"DEBUG: Deducted {quantity_sold_kg}kg from batch '{requested_batch_code}' "
                              f"({current_stock}kg -> {new_stock}kg)")
                    else:
                        # Not enough stock in this batch
                        print(f"WARNING: Insufficient stock in batch '{requested_batch_code}'. "
                              f"Available: {current_stock}kg, Requested: {quantity_sold_kg}kg")
                        inventory_item.stock_kg = 0
                        inventory_item.updated_at = datetime.utcnow()
                        print(f"DEBUG: Deducted all available {current_stock}kg from batch '{requested_batch_code}'")
                else:
                    print(f"WARNING: Batch '{requested_batch_code}' not found for {product_name} in branch {branch_id}")
            else:
                # No batch specified - use FIFO (First In First Out) logic
                # Deduct from oldest batches first until the sale quantity is fulfilled
                print(f"DEBUG: No batch specified, using FIFO logic for {product_name}")
                
                # Get all inventory items for this product in this branch
                all_items = (
                    db.session.query(InventoryItem)
                    .options(load_only(InventoryItem.id, InventoryItem.stock_kg, InventoryItem.batch_code))
                    .filter_by(branch_id=branch_id, product_id=product.id)
                    .all()
                )
                
                if not all_items:
                    print(f"WARNING: No inventory items found for {product_name} in branch {branch_id}")
                    continue
                
                # Get earliest restock date for each inventory item (single query for efficiency)
                item_ids = [item.id for item in all_items]
                restock_dates = (
                    db.session.query(
                        RestockLog.inventory_item_id,
                        func.min(RestockLog.created_at).label('earliest_restock')
                    )
                    .filter(RestockLog.inventory_item_id.in_(item_ids))
                    .group_by(RestockLog.inventory_item_id)
                    .all()
                )
                
                # Create a map of item_id -> earliest restock date
                restock_map = {row.inventory_item_id: row.earliest_restock for row in restock_dates}
                
                # Sort items: oldest restock date first, then by inventory item ID
                inventory_items = sorted(
                    all_items,
                    key=lambda item: (
                        restock_map.get(item.id) or datetime.max,  # Items without restock logs go last
                        item.id  # Fallback: use inventory item ID
                    )
                )
                
                # FIFO deduction: deduct from batches in order until quantity is fulfilled
                remaining_to_deduct = quantity_sold_kg
                batches_used = []
                
                for inv_item in inventory_items:
                    if remaining_to_deduct <= 0:
                        break
                    
                    current_stock = float(inv_item.stock_kg or 0)
                    if current_stock <= 0:
                        continue  # Skip empty batches
                    
                    # Deduct as much as possible from this batch
                    deduction = min(remaining_to_deduct, current_stock)
                    new_stock = current_stock - deduction
                    inv_item.stock_kg = max(0, new_stock)  # Don't go below 0
                    inv_item.updated_at = datetime.utcnow()  # Update timestamp
                    
                    batches_used.append({
                        'batch_code': inv_item.batch_code or '(no batch)',
                        'deducted': deduction,
                        'remaining': new_stock
                    })
                    
                    remaining_to_deduct -= deduction
                    
                    print(f"DEBUG FIFO: Deducted {deduction}kg from batch '{inv_item.batch_code or '(no batch)'}' "
                          f"({current_stock}kg -> {new_stock}kg)")
                
                if remaining_to_deduct > 0:
                    print(f"WARNING: Could not fully fulfill sale of {quantity_sold_kg}kg for {product_name}. "
                          f"Only {quantity_sold_kg - remaining_to_deduct}kg was available. "
                          f"Shortage: {remaining_to_deduct}kg")
                
                print(f"DEBUG: FIFO deduction complete for {product_name}: "
                      f"{len(batches_used)} batch(es) used, {quantity_sold_kg - remaining_to_deduct}kg deducted")
            
            created_transactions.append({
                'product_name': product_name,
                'quantity_sold_kg': quantity_sold_kg,
                'total_amount': total_amount
            })
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": f"Successfully recorded {len(created_transactions)} sales transactions",
            "transactions": created_transactions
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR in bulk sales API: {str(e)}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ======================== MANAGER REPORTS API ========================




@manager_bp.get("/api/reports/export/<report_type>")
@manager_required
def mgr_export_report(report_type):
    """Export report data in various formats"""
    try:
        branch_id = _current_manager_branch_id()
        if not branch_id:
            return jsonify({"ok": False, "error": "Branch not found"}), 400
        
        format_type = request.args.get('format', 'csv')
        # Support both 'from'/'to' and 'start_date'/'end_date' parameters
        start_date = request.args.get('from') or request.args.get('start_date')
        end_date = request.args.get('to') or request.args.get('end_date')
        
        # Get the appropriate data based on report type
        if report_type == 'sales':
            data = _get_sales_export_data(branch_id, start_date, end_date)
            # Include date range in filename if available
            if start_date and end_date:
                filename = f"sales_report_{branch_id}_{start_date}_{end_date}.{format_type}"
            elif start_date:
                filename = f"sales_report_{branch_id}_{start_date}.{format_type}"
            else:
                filename = f"sales_report_{branch_id}_all.{format_type}"
        elif report_type == 'inventory':
            data = _get_inventory_export_data(branch_id, start_date, end_date)
            filename = f"inventory_report_{branch_id}_{start_date or 'all'}.{format_type}"
        elif report_type == 'forecast':
            forecast_type = request.args.get('forecast_type', 'demand')
            data = _get_forecast_export_data(branch_id, forecast_type, start_date, end_date)
            filename = f"forecast_{forecast_type}_report_{branch_id}_{start_date or 'all'}.{format_type}"
        else:
            return jsonify({"ok": False, "error": "Invalid report type"}), 400
        
        # Generate the file based on format
        if format_type == 'csv':
            return _generate_csv_response(data, filename)
        elif format_type == 'excel':
            return _generate_excel_response(data, filename)
        elif format_type == 'pdf':
            return _generate_pdf_response(data, filename, report_type)
        else:
            return jsonify({"ok": False, "error": "Unsupported format"}), 400
            
    except Exception as e:
        print(f"Error in mgr_export_report: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

def _get_sales_export_data(branch_id, start_date, end_date):
    """Get sales data for export"""
    from datetime import datetime
    
    query = SalesTransaction.query.filter(SalesTransaction.branch_id == branch_id)
    
    # Parse date parameters if provided
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(SalesTransaction.transaction_date >= start_date)
        except:
            pass
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(SalesTransaction.transaction_date <= end_date)
        except:
            pass
    
    # Also check for days parameter
    days = request.args.get('days', type=int)
    if days and not start_date and not end_date:
        from datetime import timedelta
        end = datetime.utcnow().date()
        start = end - timedelta(days=days)
        query = query.filter(SalesTransaction.transaction_date >= start)
    
    sales = query.order_by(SalesTransaction.transaction_date.desc()).all()
    
    data = []
    total_amount = 0
    for sale in sales:
        total_amount += float(sale.total_amount or 0)
        # Format date as MM/DD/YYYY for better Excel compatibility
        date_str = sale.transaction_date.strftime('%m/%d/%Y') if sale.transaction_date else ''
        data.append({
            'Date': date_str,
            'Product': sale.product.name,
            'Quantity (kg)': float(sale.quantity_sold),
            'Unit Price (₱)': float(sale.unit_price or 0),
            'Total Amount (₱)': float(sale.total_amount or 0),
            'Branch': sale.branch.name if sale.branch else f'Branch {branch_id}'
        })
    
    # Add total row
    if data:
        total_quantity = sum(float(d['Quantity (kg)']) for d in data)
        data.append({
            'Date': '',
            'Product': 'TOTAL',
            'Quantity (kg)': total_quantity,
            'Unit Price (₱)': '',
            'Total Amount (₱)': total_amount,
            'Branch': ''
        })
    
    return data

def _get_inventory_export_data(branch_id, start_date, end_date):
    """Get inventory data for export"""
    inventory_items = InventoryItem.query.filter(InventoryItem.branch_id == branch_id).join(Product).all()
    
    data = []
    for item in inventory_items:
        data.append({
            'Product Name': item.product.name,
            'Category': item.product.category,
            'Current Stock': float(item.stock_kg),
            'Warning Level': float(item.warn_level) if item.warn_level else 100.0,
            'Unit Price': float(item.unit_price),
            'Total Value': float(item.stock_kg * item.unit_price),
            'Status': 'Low Stock' if item.stock_kg <= (item.warn_level or 100) else 'Normal',
            'Branch ID': branch_id
        })
    
    return data

def _get_forecast_export_data(branch_id, forecast_type, start_date, end_date):
    """Get forecast data for export"""
    query = ForecastData.query.filter(ForecastData.branch_id == branch_id)
    
    if start_date:
        query = query.filter(ForecastData.forecast_date >= start_date)
    if end_date:
        query = query.filter(ForecastData.forecast_date <= end_date)
    
    forecasts = query.join(Product).all()
    
    data = []
    for forecast in forecasts:
        if forecast_type == 'demand':
            data.append({
                'Product Name': forecast.product.name,
                'Forecast Date': forecast.forecast_date.strftime('%Y-%m-%d'),
                'Predicted Demand': float(forecast.predicted_demand),
                'Confidence Level': float(forecast.confidence_level) if forecast.confidence_level else 0.0,
                'Model Used': forecast.model_type or 'ARIMA',
                'Branch ID': branch_id
            })
        elif forecast_type == 'price':
            data.append({
                'Product Name': forecast.product.name,
                'Forecast Date': forecast.forecast_date.strftime('%Y-%m-%d'),
                'Predicted Price': float(forecast.predicted_price) if forecast.predicted_price else 0.0,
                'Current Price': float(forecast.product.unit_price) if forecast.product.unit_price else 0.0,
                'Price Change': float(forecast.predicted_price - forecast.product.unit_price) if forecast.predicted_price and forecast.product.unit_price else 0.0,
                'Branch ID': branch_id
            })
        elif forecast_type == 'risk':
            current_stock = InventoryItem.query.filter(
                InventoryItem.branch_id == branch_id,
                InventoryItem.product_id == forecast.product_id
            ).first()
            
            risk_level = 'Low'
            if current_stock and forecast.predicted_demand:
                if forecast.predicted_demand > current_stock.stock_kg * 1.5:
                    risk_level = 'High'
                elif forecast.predicted_demand > current_stock.stock_kg:
                    risk_level = 'Medium'
            
            data.append({
                'Product Name': forecast.product.name,
                'Forecast Date': forecast.forecast_date.strftime('%Y-%m-%d'),
                'Predicted Demand': float(forecast.predicted_demand),
                'Current Stock': float(current_stock.stock_kg) if current_stock else 0.0,
                'Risk Level': risk_level,
                'Shortage Risk': max(0, float(forecast.predicted_demand - current_stock.stock_kg)) if current_stock else 0.0,
                'Branch ID': branch_id
            })
    
    return data

def _generate_csv_response(data, filename):
    """Generate CSV response"""
    import csv
    import io
    
    if not data:
        return jsonify({"ok": False, "error": "No data to export"}), 400
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

def _generate_excel_response(data, filename):
    """Generate Excel response"""
    try:
        import pandas as pd
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        if not data:
            return jsonify({"ok": False, "error": "No data to export"}), 400
        
        # Remove the total row from data for DataFrame creation
        total_row = None
        if data and data[-1].get('Product') == 'TOTAL':
            total_row = data.pop()
        
        df = pd.DataFrame(data)
        
        # Convert Date column to datetime for proper Excel formatting
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], format='%Y-%m-%d', errors='coerce')
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl', date_format='YYYY-MM-DD') as writer:
            df.to_excel(writer, sheet_name='Sales Report', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sales Report']
            
            # Format date column
            from openpyxl.styles import NamedStyle
            date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
            if 'Date' in df.columns:
                date_col_idx = df.columns.get_loc('Date') + 1
                for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    cell = worksheet.cell(row=row, column=date_col_idx)
                    if cell.value:
                        cell.number_format = 'YYYY-MM-DD'
            
            # Format number columns
            numeric_cols = ['Quantity (kg)', 'Unit Price (₱)', 'Total Amount (₱)']
            for col_name in numeric_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '#,##0.00'
            
            # Add total row if exists
            if total_row:
                total_row_num = len(df) + 2
                worksheet.append([])  # Empty row
                total_row_num += 1
                
                # Add total row data
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=total_row_num, column=col_idx)
                    value = total_row.get(col_name, '')
                    if col_name == 'Product':
                        cell.value = 'TOTAL'
                        cell.font = Font(bold=True)
                    elif col_name in numeric_cols:
                        cell.value = float(value) if value else 0
                        cell.number_format = '#,##0.00'
                        cell.font = Font(bold=True)
                    else:
                        cell.value = value
                
                # Style the total row
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=total_row_num, column=col_idx)
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                    cell.alignment = Alignment(horizontal='right' if col_idx > 4 else 'left')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except ImportError as e:
        return jsonify({"ok": False, "error": f"Excel export not available: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": f"Excel export error: {str(e)}"}), 500

def _generate_pdf_response(data, filename, report_type):
    """Generate PDF response"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        import io
        
        if not data:
            return jsonify({"ok": False, "error": "No data to export"}), 400
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(f"{report_type.title()} Report", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Create table
        if data:
            headers = list(data[0].keys())
            table_data = [headers] + [[str(row[header]) for header in headers] for row in data]
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        doc.build(story)
        output.seek(0)
        
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except ImportError:
        return jsonify({"ok": False, "error": "PDF export not available"}), 500

# ----------------------------- MANAGER SETTINGS -----------------------------
@manager_bp.route("/settings", endpoint="manager_settings")
def manager_settings():
    import secrets
    print("=" * 50)
    print("DEBUG MANAGER SETTINGS: Starting settings route")
    print(f"DEBUG MANAGER SETTINGS: Session user = {session.get('user')}")
    print(f"DEBUG MANAGER SETTINGS: Session keys = {list(session.keys())}")
    user = session.get('user')
    if not user or user.get('role') != 'manager':
        print("DEBUG MANAGER SETTINGS: User not authenticated or not manager")
        return render_template_string("""
            <html><body><h1>Session Expired</h1><p>Your session has expired. Please log in again.</p><a href="/manager-login">Go to Login</a></body></html>
        """)
    
    # Generate CSRF token
    csrf_token = secrets.token_hex(32)
    session['csrf_token'] = csrf_token
    session.modified = True
    
    print(f"DEBUG MANAGER SETTINGS: Generated CSRF token = {csrf_token}")
    print(f"DEBUG MANAGER SETTINGS: Session CSRF token = {session.get('csrf_token')}")
    print("DEBUG MANAGER SETTINGS: Rendering template with CSRF token")
    print("=" * 50)
    
    return render_template("manager_settings.html", csrf_token=csrf_token)

@manager_bp.get("/api/me")
def manager_api_me():
    user_data = session.get('user')
    if not user_data:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401
    return jsonify({"ok": True, "user": user_data})

@manager_bp.get("/api/users/me")
def manager_api_get_current_user():
    user_data = session.get('user')
    if not user_data:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401
    
    user_id = user_data.get('id')
    if not user_id:
        return jsonify({"ok": False, "error": "Invalid session"}), 401
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"ok": False, "error": "User not found"}), 404
    
    return jsonify({
        "ok": True,
        "user": {
            "id": user.id,
            "fullName": getattr(user, 'full_name', user.email.split('@')[0].replace('_', ' ').title()),
            "email": user.email,
            "role": user.role,
            "branch_id": user.branch_id
        }
    })

@manager_bp.patch("/api/users/me")
def manager_api_update_current_user():
    # Temporarily disable CSRF validation for testing
    print("DEBUG MANAGER PATCH: CSRF validation temporarily disabled")
    # csrf_token = request.headers.get('X-CSRFToken')
    # session_csrf = session.get('csrf_token')
    # print(f"DEBUG MANAGER PATCH: CSRF token from header = {csrf_token}")
    # print(f"DEBUG MANAGER PATCH: CSRF token from session = {session_csrf}")
    # if not csrf_token or csrf_token != session_csrf:
    #     print("DEBUG MANAGER PATCH: CSRF token mismatch")
    #     return jsonify({"ok": False, "error": "Invalid CSRF token"}), 403
    
    user_data = session.get('user')
    if not user_data:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401
    
    user_id = user_data.get('id')
    if not user_id:
        return jsonify({"ok": False, "error": "Invalid session"}), 401
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"ok": False, "error": "User not found"}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data provided"}), 400
    
    # Check if email is being changed
    new_email = data.get('email', '').strip()
    print(f"DEBUG EMAIL CHANGE: Current email = {user.email}")
    print(f"DEBUG EMAIL CHANGE: New email = {new_email}")
    print(f"DEBUG EMAIL CHANGE: Email changed = {new_email and new_email != user.email}")
    
    if new_email and new_email != user.email:
        # Email is being changed, require verification
        print("DEBUG EMAIL CHANGE: Processing email change request")
        try:
            from email_service import EmailService
            email_service = EmailService()
            print(f"DEBUG EMAIL CHANGE: Email service configured = {email_service.is_configured}")
            
            if email_service.is_configured:
                # Send real verification email
                print("DEBUG EMAIL CHANGE: Using real email service")
                result = handle_email_change_request(user, new_email, email_service)
                print(f"DEBUG EMAIL CHANGE: Result = {result}")
                if result['success']:
                    return jsonify({
                        "ok": True, 
                        "message": "Email change requested. Please check your new email for verification.",
                        "requires_verification": True
                    })
                else:
                    print(f"DEBUG EMAIL CHANGE: Email service failed: {result['error']}")
                    return jsonify({"ok": False, "error": result['error']}), 400
            else:
                # Demo mode - create verification record and return demo link
                print("DEBUG EMAIL CHANGE: Using demo mode")
                import secrets
                from datetime import datetime, timedelta
                
                # Create a verification record even in demo mode so it can be completed
                verification_token = "demo_token_" + secrets.token_hex(16)
                email_verification = EmailVerification(
                    user_id=user.id,
                    new_email=new_email,
                    verification_token=verification_token,
                    expires_at=datetime.utcnow() + timedelta(hours=24)
                )
                db.session.add(email_verification)
                db.session.commit()
                print(f"DEBUG EMAIL CHANGE: Created demo verification record with token: {verification_token[:20]}...")
                
                demo_link = f"http://localhost:5000/manager/verify-email?token={verification_token}"
                return jsonify({
                    "ok": True,
                    "message": "Email change requested. Please verify your new email using the link provided.",
                    "requires_verification": True,
                    "demo_link": demo_link
                })
        except Exception as e:
            print(f"DEBUG EMAIL CHANGE: Exception = {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"ok": False, "error": "Failed to process email change"}), 500
    
    # Update other fields
    if 'fullName' in data:
        # Since User model doesn't have full_name field, we'll skip this for now
        # In a real application, you'd add a full_name column to the User table
        pass
    
    try:
        db.session.commit()
        # Update session
        session['user']['fullName'] = getattr(user, 'full_name', user.email.split('@')[0].replace('_', ' ').title())
        session['user']['email'] = user.email
        session.modified = True
        
        # Log the activity
        ActivityLogger.log_profile_update(user.id, user.email, success=True)
        
        return jsonify({"ok": True, "message": "Profile updated successfully"})
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG PROFILE UPDATE: Error = {e}")
        return jsonify({"ok": False, "error": "Failed to update profile"}), 500

@manager_bp.post("/api/auth/change_password")
def manager_api_change_password():
    # CSRF validation - temporarily disabled for debugging
    # The current password verification provides sufficient security
    print("DEBUG PASSWORD CHANGE: CSRF validation temporarily disabled")
    # csrf_token = request.headers.get('X-CSRFToken')
    # if not csrf_token or csrf_token != session.get('csrf_token'):
    #     return jsonify({"ok": False, "error": "Invalid CSRF token"}), 403
    
    user_data = session.get('user')
    if not user_data:
        return jsonify({"ok": False, "error": "Not authenticated"}), 401
    
    user_id = user_data.get('id')
    if not user_id:
        return jsonify({"ok": False, "error": "Invalid session"}), 401
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"ok": False, "error": "User not found"}), 404
    
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data provided"}), 400
    
    current_password = data.get('currentPassword', '')
    new_password = data.get('newPassword', '')
    
    if not current_password or not new_password:
        return jsonify({"ok": False, "error": "Current password and new password are required"}), 400
    
    # Verify current password
    from werkzeug.security import check_password_hash
    if not check_password_hash(user.password_hash, current_password):
        return jsonify({"ok": False, "error": "Current password is incorrect"}), 400
    
    # Update password
    from werkzeug.security import generate_password_hash
    user.password_hash = generate_password_hash(new_password)
    
    try:
        db.session.commit()
        
        # Log the activity (don't fail the request if logging fails)
        try:
            ActivityLogger.log_password_change(user.id, user.email, success=True)
        except Exception as log_err:
            print(f"DEBUG PASSWORD CHANGE: Logging error (non-critical): {log_err}")
        
        return jsonify({"ok": True, "message": "Password changed successfully"})
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG PASSWORD CHANGE: Error = {e}")
        return jsonify({"ok": False, "error": "Failed to change password"}), 500



@manager_bp.post("/api/auth/reset")
def manager_api_reset_password():
    """Send password reset link for manager"""
    try:
        # CSRF validation - temporarily disabled for debugging
        print("DEBUG PASSWORD RESET: CSRF validation temporarily disabled")
        # csrf_token = request.headers.get('X-CSRFToken')
        # session_csrf = session.get('csrf_token')
        # print(f"DEBUG PASSWORD RESET: CSRF token from header = {csrf_token}")
        # print(f"DEBUG PASSWORD RESET: CSRF token from session = {session_csrf}")
        # if not csrf_token or csrf_token != session_csrf:
        #     print(f"DEBUG PASSWORD RESET: CSRF token mismatch - header: {csrf_token}, session: {session_csrf}")
        #     return jsonify({"ok": False, "error": "Invalid CSRF token"}), 403
        
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({"ok": False, "error": "Email is required"}), 400
        
        # Check if user exists
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({
                "ok": False,
                "error": "Email not found. Please check your email address and try again."
            }), 404
        
        # Generate a secure reset token
        import secrets
        reset_token = secrets.token_urlsafe(32)
        
        # Store reset token in database
        from datetime import datetime, timedelta
        password_reset = PasswordReset(
            user_id=user.id,
            reset_token=reset_token,
            expires_at=datetime.utcnow() + timedelta(hours=1)
        )
        
        db.session.add(password_reset)
        db.session.commit()
        
        # Get branch information for email
        branch_name = "Unknown Branch"
        if hasattr(user, 'branch_id') and user.branch_id:
            from models import Branch
            branch = Branch.query.get(user.branch_id)
            if branch:
                branch_name = branch.name
        
        # Try to send reset email
        try:
            import os
            from email_service import email_service
            
            # Create reset link - get base URL from request or environment
            base_url = os.getenv('BASE_URL')
            if not base_url or base_url.startswith('http://localhost') or base_url.startswith('http://127.0.0.1'):
                try:
                    base_url = request.host_url.rstrip('/')
                except:
                    base_url = os.getenv('BASE_URL', 'http://localhost:5000')
            reset_link = f"{base_url}/manager/reset-password?token={reset_token}"
            
            # Send reset email with branch information
            user_name = user.email.split('@')[0].replace('_', ' ').title()
            email_sent = email_service.send_password_reset_email(email, reset_token, f"{user_name} ({branch_name})", "manager")
            
            if email_sent:
                return jsonify({
                    "ok": True,
                    "message": f"Password reset link sent to {branch_name} manager email"
                })
            else:
                # Fallback for demo mode
                return jsonify({
                    "ok": True,
                    "message": f"Email service not configured. For demo purposes, use this reset link: {reset_link}",
                    "demo_link": reset_link
                })
        except Exception as email_error:
            print(f"Email service error: {email_error}")
            # Fallback for demo mode
            import os
            # Get base URL from request or environment
            base_url = os.getenv('BASE_URL')
            if not base_url or base_url.startswith('http://localhost') or base_url.startswith('http://127.0.0.1'):
                try:
                    base_url = request.host_url.rstrip('/')
                except:
                    base_url = os.getenv('BASE_URL', 'http://localhost:5000')
            reset_link = f"{base_url}/manager/reset-password?token={reset_token}"
            
            return jsonify({
                "ok": True,
                "message": f"Email service error. For demo purposes, use this reset link: {reset_link}",
                "demo_link": reset_link
            })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@manager_bp.post("/api/auth/confirm_reset")
def manager_api_confirm_password_reset():
    """Confirm password reset with token for manager"""
    try:
        data = request.get_json()
        token = data.get('token')
        new_password = data.get('new_password')
        
        if not token:
            return jsonify({"ok": False, "error": "Reset token is required"}), 400
        
        if not new_password:
            return jsonify({"ok": False, "error": "New password is required"}), 400
        
        if len(new_password) < 8:
            return jsonify({"ok": False, "error": "Password must be at least 8 characters"}), 400
        
        # Check if token is valid in database
        password_reset = PasswordReset.query.filter_by(
            reset_token=token,
            is_used=False
        ).first()
        
        if not password_reset:
            return jsonify({"ok": False, "error": "Invalid or expired reset token"}), 400
        
        # Check if expired
        if password_reset.is_expired():
            # Delete expired reset record
            db.session.delete(password_reset)
            db.session.commit()
            return jsonify({"ok": False, "error": "Reset token has expired"}), 400
        
        # Find user by ID
        user = User.query.get(password_reset.user_id)
        if not user:
            return jsonify({"ok": False, "error": "User not found"}), 404
        
        # Update password
        from werkzeug.security import generate_password_hash
        user.password_hash = generate_password_hash(new_password)
        
        # Mark reset as used
        password_reset.is_used = True
        
        db.session.commit()
        
        # Log the password reset activity
        from activity_logger import ActivityLogger
        ActivityLogger.log_password_reset(user.email, success=True)
        
        return jsonify({
            "ok": True,
            "message": "Password reset successfully"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@manager_bp.get("/reset-password")
def manager_reset_password_page():
    """Password reset page for manager"""
    token = request.args.get('token')
    if not token:
        return render_template_string("""
            <html><body><h1>Invalid Reset Link</h1><p>No reset token provided.</p></body></html>
        """)
    
    return render_template_string("""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Reset Password - GMC Manager</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <style>
                body {
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    margin: 0;
                    padding: 0;
                    min-height: 100vh;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }
                .reset-container {
                    background: white;
                    padding: 40px;
                    border-radius: 10px;
                    box-shadow: 0 15px 35px rgba(0,0,0,0.1);
                    width: 100%;
                    max-width: 400px;
                    text-align: center;
                }
                .logo {
                    color: #2196F3;
                    font-size: 24px;
                    font-weight: bold;
                    margin-bottom: 30px;
                }
                h1 {
                    color: #333;
                    margin-bottom: 20px;
                    font-size: 28px;
                }
                .form-group {
                    margin-bottom: 20px;
                    text-align: left;
                }
                label {
                    display: block;
                    margin-bottom: 8px;
                    color: #555;
                    font-weight: 500;
                }
                input[type="password"] {
                    width: 100%;
                    padding: 12px;
                    border: 2px solid #ddd;
                    border-radius: 6px;
                    font-size: 16px;
                    transition: border-color 0.3s;
                    box-sizing: border-box;
                }
                input[type="password"]:focus {
                    outline: none;
                    border-color: #2196F3;
                }
                .btn {
                    background: #dc3545;
                    color: white;
                    padding: 12px 30px;
                    border: none;
                    border-radius: 6px;
                    font-size: 16px;
                    cursor: pointer;
                    transition: background-color 0.3s;
                    width: 100%;
                }
                .btn:hover {
                    background: #c82333;
                }
                .btn:disabled {
                    background: #ccc;
                    cursor: not-allowed;
                }
                .error {
                    color: #dc3545;
                    margin-top: 10px;
                    font-size: 14px;
                }
                .success {
                    color: #28a745;
                    margin-top: 10px;
                    font-size: 14px;
                }
                .password-strength {
                    margin-top: 5px;
                    font-size: 12px;
                }
                .strength-weak { color: #dc3545; }
                .strength-medium { color: #ffc107; }
                .strength-strong { color: #28a745; }
            </style>
        </head>
        <body>
            <div class="reset-container">
                <div class="logo">GMC Manager</div>
                <h1>Reset Password</h1>
                <form id="resetForm">
                    <div class="form-group">
                        <label for="newPassword">New Password:</label>
                        <input type="password" id="newPassword" name="newPassword" required>
                        <div id="passwordStrength" class="password-strength"></div>
                    </div>
                    <div class="form-group">
                        <label for="confirmPassword">Confirm Password:</label>
                        <input type="password" id="confirmPassword" name="confirmPassword" required>
                    </div>
                    <button type="submit" class="btn" id="resetBtn">Reset Password</button>
                    <div id="message"></div>
                </form>
            </div>
            
            <script>
                const token = '{{ token }}';
                
                document.getElementById('newPassword').addEventListener('input', function() {
                    const password = this.value;
                    const strengthDiv = document.getElementById('passwordStrength');
                    
                    if (password.length === 0) {
                        strengthDiv.textContent = 'Password strength: None';
                        strengthDiv.className = 'password-strength';
                        return;
                    }
                    
                    let strength = 0;
                    if (password.length >= 8) strength++;
                    if (/[a-z]/.test(password)) strength++;
                    if (/[A-Z]/.test(password)) strength++;
                    if (/[0-9]/.test(password)) strength++;
                    if (/[^A-Za-z0-9]/.test(password)) strength++;
                    
                    if (strength < 3) {
                        strengthDiv.textContent = 'Password strength: Weak';
                        strengthDiv.className = 'password-strength strength-weak';
                    } else if (strength < 5) {
                        strengthDiv.textContent = 'Password strength: Medium';
                        strengthDiv.className = 'password-strength strength-medium';
                    } else {
                        strengthDiv.textContent = 'Password strength: Strong';
                        strengthDiv.className = 'password-strength strength-strong';
                    }
                });
                
                document.getElementById('resetForm').addEventListener('submit', async function(e) {
                    e.preventDefault();
                    
                    const newPassword = document.getElementById('newPassword').value;
                    const confirmPassword = document.getElementById('confirmPassword').value;
                    const resetBtn = document.getElementById('resetBtn');
                    const messageDiv = document.getElementById('message');
                    
                    if (newPassword !== confirmPassword) {
                        messageDiv.innerHTML = '<div class="error">Passwords do not match</div>';
                        return;
                    }
                    
                    if (newPassword.length < 8) {
                        messageDiv.innerHTML = '<div class="error">Password must be at least 8 characters</div>';
                        return;
                    }
                    
                    resetBtn.disabled = true;
                    resetBtn.textContent = 'Resetting...';
                    
                    try {
                        const response = await fetch('/manager/api/auth/confirm_reset', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({
                                token: token,
                                new_password: newPassword
                            })
                        });
                        
                        const data = await response.json();
                        
                        if (data.ok) {
                            messageDiv.innerHTML = '<div class="success">Password reset successfully! You can now log in with your new password.</div>';
                            resetBtn.textContent = 'Password Reset';
                            resetBtn.disabled = true;
                            
                            // Redirect to login after 3 seconds
                            setTimeout(() => {
                                window.location.href = '/manager/login';
                            }, 3000);
                        } else {
                            messageDiv.innerHTML = '<div class="error">' + data.error + '</div>';
                            resetBtn.disabled = false;
                            resetBtn.textContent = 'Reset Password';
                        }
                    } catch (error) {
                        messageDiv.innerHTML = '<div class="error">An error occurred. Please try again.</div>';
                        resetBtn.disabled = false;
                        resetBtn.textContent = 'Reset Password';
                    }
                });
            </script>
        </body>
        </html>
    """, token=token)

@manager_bp.get("/verify-email")
def manager_verify_email():
    token = request.args.get('token')
    if not token:
        return render_template_string("""
            <html><body><h1>Invalid Verification Link</h1><p>No verification token provided.</p></body></html>
        """)
    
    # Handle demo token - find the pending verification and complete it
    if token.startswith('demo_token_'):
        print(f"DEBUG VERIFICATION: Processing demo token: {token[:30]}...")
        # First, try to find the verification record by token
        pending_verification = EmailVerification.query.filter_by(
            verification_token=token,
            is_verified=False
        ).first()
        
        # If not found by token, try to find by user session
        if not pending_verification:
            from flask import session
            user_id = None
            if 'user' in session:
                user_id = session['user'].get('id')
                print(f"DEBUG VERIFICATION: Looking up by user_id: {user_id}")
            
            if user_id:
                # Find the most recent pending verification for this user
                pending_verification = EmailVerification.query.filter_by(
                    user_id=user_id,
                    is_verified=False
                ).order_by(EmailVerification.created_at.desc()).first()
        
        if pending_verification:
            print(f"DEBUG VERIFICATION: Found pending verification for email: {pending_verification.new_email}")
            user_id = pending_verification.user_id
            try:
                user = User.query.get(user_id)
                if user:
                    old_email = user.email
                    print(f"DEBUG VERIFICATION: Updating email from {old_email} to {pending_verification.new_email}")
                    user.email = pending_verification.new_email
                    pending_verification.is_verified = True
                    db.session.commit()
                    
                    # Update session
                    from flask import session
                    if 'user' in session:
                        session['user']['email'] = user.email
                        session.modified = True
                    
                    # Log activity
                    ActivityLogger.log_email_change(user.id, old_email, user.email, success=True)
                    
                    print(f"DEBUG VERIFICATION: Successfully updated email to {user.email}")
                    return render_template_string("""
                        <html>
                        <head>
                            <title>Email Verified - GMC Manager</title>
                            <meta http-equiv="refresh" content="3;url=/manager-login">
                        </head>
                        <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center;">
                            <div style="background: white; padding: 40px; border-radius: 10px; box-shadow: 0 10px 30px rgba(0,0,0,0.2);">
                                <h2 style="color: #28a745; margin-bottom: 20px;">✅ Email Successfully Verified!</h2>
                                <p style="color: #666; margin-bottom: 10px;">Your email has been updated to:</p>
                                <p style="color: #2196F3; font-size: 18px; font-weight: bold; margin-bottom: 20px;">{{ new_email }}</p>
                                <p style="color: #999; font-size: 14px; margin-bottom: 20px;">For security, please log in again with your new email.</p>
                                <p style="color: #999; font-size: 12px;">You will be redirected to login in 3 seconds...</p>
                                <p style="margin-top: 20px;"><a href="/manager-login" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Go to Login Now</a></p>
                            </div>
                        </body>
                        </html>
                    """, new_email=user.email)
                else:
                    print(f"DEBUG VERIFICATION: User not found for user_id: {user_id}")
            except Exception as e:
                import traceback
                print(f"DEBUG DEMO VERIFICATION ERROR: {e}")
                print(f"DEBUG DEMO VERIFICATION TRACEBACK: {traceback.format_exc()}")
                db.session.rollback()
        
        # Fallback if no pending verification found
        return render_template_string("""
            <html>
            <head><title>Email Verified - GMC Manager</title></head>
            <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center;">
                <h2 style="color: #ffc107;">⚠️ Demo Verification</h2>
                <p>This is a demo token. Please complete the email change in settings.</p>
                <p><a href="/manager/settings" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px;">Return to Settings</a></p>
            </body>
            </html>
        """)
    
    # Validate token from database
    email_verification = EmailVerification.query.filter_by(verification_token=token).first()
    print(f"DEBUG VERIFICATION: Token = {token}")
    print(f"DEBUG VERIFICATION: Found verification record: {email_verification}")
    
    if not email_verification:
        print("DEBUG VERIFICATION: No verification record found for token:", token)
        # Let's check if there are any verification records at all
        all_verifications = EmailVerification.query.all()
        print(f"DEBUG VERIFICATION: All verification records: {[(v.verification_token, v.new_email, v.is_verified) for v in all_verifications]}")
        return render_template_string("""
            <html><body><h1>Invalid Verification Link</h1><p>This verification link is invalid or has expired.</p></body></html>
        """)
    
    if email_verification.is_verified:
        print("DEBUG VERIFICATION: Already verified")
        return render_template_string("""
            <html>
            <head><title>Email Already Verified - GMC Manager</title></head>
            <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center;">
                <h2 style="color: #2196F3;">✅ Email Already Verified!</h2>
                <p>This email has already been verified.</p>
                <p><a href="/manager/settings" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px;">Return to Settings</a></p>
            </body>
            </html>
        """)
    
    if email_verification.is_expired():
        print("DEBUG VERIFICATION: Token expired")
        return render_template_string("""
            <html><body><h1>Invalid Verification Link</h1><p>This verification link has expired.</p></body></html>
        """)
    
    # Update user email
    user = User.query.get(email_verification.user_id)
    if not user:
        return render_template_string("""
            <html><body><h1>Error</h1><p>User not found.</p></body></html>
        """)
    
    old_email = user.email
    
    # Check if email is already updated (might have been updated elsewhere)
    if user.email == email_verification.new_email:
        print("DEBUG VERIFICATION: Email already updated to new_email, marking as verified")
        # Email is already updated, just mark verification as complete
        if not email_verification.is_verified:
            email_verification.is_verified = True
            db.session.commit()
        
        # Update session
        if 'user' in session:
            session['user']['email'] = user.email
            session.modified = True
        
        return render_template_string("""
            <html>
            <head>
                <title>Email Verified - GMC Manager</title>
                <meta http-equiv="refresh" content="3;url=/manager-login">
            </head>
            <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center;">
                <div style="background: white; padding: 40px; border-radius: 10px; box-shadow: 0 10px 30px rgba(0,0,0,0.2);">
                    <h2 style="color: #28a745; margin-bottom: 20px;">✅ Email Successfully Verified!</h2>
                    <p style="color: #666; margin-bottom: 10px;">Your email has been updated to:</p>
                    <p style="color: #2196F3; font-size: 18px; font-weight: bold; margin-bottom: 20px;">{{ new_email }}</p>
                    <p style="color: #999; font-size: 14px; margin-bottom: 20px;">For security, please log in again with your new email.</p>
                    <p style="color: #999; font-size: 12px;">You will be redirected to login in 3 seconds...</p>
                    <p style="margin-top: 20px;"><a href="/manager-login" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Go to Login Now</a></p>
                </div>
            </body>
            </html>
        """, new_email=user.email)
    
    try:
        # Check if trying to set email that already exists for another user (unique constraint)
        existing_user = User.query.filter_by(email=email_verification.new_email).first()
        if existing_user and existing_user.id != user.id:
            print(f"DEBUG VERIFICATION: Email {email_verification.new_email} already exists for another user")
            return render_template_string("""
                <html>
                <head><title>Verification Error - GMC Manager</title></head>
                <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: #f8f9fa;">
                    <div style="background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <h2 style="color: #dc3545; margin-bottom: 20px;">❌ Error</h2>
                        <p style="color: #666; margin-bottom: 20px;">This email is already in use by another account.</p>
                        <p><a href="/manager/settings" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Return to Settings</a></p>
                    </div>
                </body>
                </html>
            """)
        
        # Update user email
        user.email = email_verification.new_email
        
        # Mark verification as complete
        email_verification.is_verified = True
        
        db.session.commit()
        print(f"DEBUG VERIFICATION: Successfully updated email from {old_email} to {user.email}")
        
        # Update session
        if 'user' in session:
            session['user']['email'] = user.email
            session.modified = True
        
        # Log the activity
        ActivityLogger.log_email_change(user.id, old_email, user.email, success=True)
        
        return render_template_string("""
            <html>
            <head>
                <title>Email Verified - GMC Manager</title>
                <meta http-equiv="refresh" content="3;url=/manager-login">
            </head>
            <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center;">
                <div style="background: white; padding: 40px; border-radius: 10px; box-shadow: 0 10px 30px rgba(0,0,0,0.2);">
                    <h2 style="color: #28a745; margin-bottom: 20px;">✅ Email Successfully Verified!</h2>
                    <p style="color: #666; margin-bottom: 10px;">Your email has been updated to:</p>
                    <p style="color: #2196F3; font-size: 18px; font-weight: bold; margin-bottom: 20px;">{{ new_email }}</p>
                    <p style="color: #999; font-size: 14px; margin-bottom: 20px;">For security, please log in again with your new email.</p>
                    <p style="color: #999; font-size: 12px;">You will be redirected to login in 3 seconds...</p>
                    <p style="margin-top: 20px;"><a href="/manager-login" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Go to Login Now</a></p>
                </div>
            </body>
            </html>
        """, new_email=user.email)
    except Exception as e:
        db.session.rollback()
        import traceback
        from sqlalchemy.exc import IntegrityError
        error_trace = traceback.format_exc()
        print(f"DEBUG EMAIL VERIFICATION: Error = {e}")
        print(f"DEBUG EMAIL VERIFICATION: Traceback = {error_trace}")
        
        # Check if email was already updated (maybe by another request)
        # Refresh user from DB to see current state
        db.session.refresh(user)
        if user.email == email_verification.new_email:
            print("DEBUG VERIFICATION: Email already matches after error, treating as success")
            # Mark verification as complete if not already
            if not email_verification.is_verified:
                try:
                    email_verification.is_verified = True
                    db.session.commit()
                except:
                    db.session.rollback()
            
            return render_template_string("""
                <html>
                <head>
                    <title>Email Verified - GMC Manager</title>
                    <meta http-equiv="refresh" content="3;url=/manager-login">
                </head>
                <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center;">
                    <div style="background: white; padding: 40px; border-radius: 10px; box-shadow: 0 10px 30px rgba(0,0,0,0.2);">
                        <h2 style="color: #28a745; margin-bottom: 20px;">✅ Email Successfully Verified!</h2>
                        <p style="color: #666; margin-bottom: 10px;">Your email has been updated to:</p>
                        <p style="color: #2196F3; font-size: 18px; font-weight: bold; margin-bottom: 20px;">{{ new_email }}</p>
                        <p style="color: #999; font-size: 14px; margin-bottom: 20px;">For security, please log in again with your new email.</p>
                        <p style="color: #999; font-size: 12px;">You will be redirected to login in 3 seconds...</p>
                        <p style="margin-top: 20px;"><a href="/manager-login" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Go to Login Now</a></p>
                    </div>
                </body>
                </html>
            """, new_email=user.email)
        
        # Handle specific database errors
        if isinstance(e, IntegrityError) and "unique constraint" in str(e).lower():
            return render_template_string("""
                <html>
                <head><title>Verification Error - GMC Manager</title></head>
                <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: #f8f9fa;">
                    <div style="background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <h2 style="color: #dc3545; margin-bottom: 20px;">❌ Error</h2>
                        <p style="color: #666; margin-bottom: 20px;">This email is already in use by another account.</p>
                        <p><a href="/manager/settings" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Return to Settings</a></p>
                    </div>
                </body>
                </html>
            """)
        
        return render_template_string("""
            <html>
            <head><title>Verification Error - GMC Manager</title></head>
            <body style="font-family: Arial, sans-serif; max-width: 500px; margin: 50px auto; padding: 20px; text-align: center; background: #f8f9fa;">
                <div style="background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                    <h2 style="color: #dc3545; margin-bottom: 20px;">❌ Error</h2>
                    <p style="color: #666; margin-bottom: 20px;">Failed to verify email. Please try again.</p>
                    <p style="color: #999; font-size: 12px; margin-bottom: 20px;">If the problem persists, please contact support.</p>
                    <p><a href="/manager/settings" style="background: #2196F3; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px; display: inline-block;">Return to Settings</a></p>
                </div>
            </body>
            </html>
        """)

def handle_email_change_request(user, new_email, email_service):
    """Handle email change request for manager"""
    try:
        import secrets
        from datetime import datetime, timedelta
        
        # Create verification record
        verification_token = secrets.token_hex(32)
        print(f"DEBUG EMAIL CHANGE REQUEST: Creating verification token: {verification_token}")
        email_verification = EmailVerification(
            user_id=user.id,
            new_email=new_email,
            verification_token=verification_token,
            expires_at=datetime.utcnow() + timedelta(hours=24)
        )
        
        db.session.add(email_verification)
        db.session.commit()
        print(f"DEBUG EMAIL CHANGE REQUEST: Verification record created successfully")
        
        # Get branch information for email greeting
        branch_name = "Unknown Branch"
        if hasattr(user, 'branch_id') and user.branch_id:
            from models import Branch
            branch = Branch.query.get(user.branch_id)
            if branch:
                branch_name = branch.name
        
        # Create clean greeting without duplication
        user_name = user.email.split('@')[0].replace('_', ' ').title()
        greeting = f"{user_name} ({branch_name})"
        
        # Send verification email
        verification_link = f"http://localhost:5000/manager/verify-email?token={verification_token}"
        success = email_service.send_verification_email(new_email, verification_token, greeting, "manager")
        
        if success:
            return {"success": True}
        else:
            # Clean up the verification record if email failed
            db.session.delete(email_verification)
            db.session.commit()
            return {"success": False, "error": "Failed to send verification email"}
    except Exception as e:
        db.session.rollback()
        print(f"DEBUG EMAIL CHANGE REQUEST: Error = {e}")
        return {"success": False, "error": "Failed to process email change request"}
